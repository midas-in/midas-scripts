#!/usr/bin/env python3
"""
MIDAS Image Curation System  v2.3
Clinical Image Archival Application for Hospital Curators
  + QC Scanner: corruption · quality · resolution per modality ·
                NDPI/WSI support · histogram export ·
                Case_ID / VISIT_Date / Image_Name identifiers
  + Dataset Counter: recursive case folder scan → styled Excel workbook
                     (Institute · Case_ID · per-modality image counts)

Requirements (core):
    pip install PyQt6 Pillow numpy matplotlib openpyxl

Optional — WSI/NDPI support:
    pip install openslide-python
    (also needs the OpenSlide C library installed on your OS)
    Windows: https://openslide.org/download/
    Linux  : sudo apt install openslide-tools
    macOS  : brew install openslide

Usage:
    python midas_curation_v2_3.py
"""

import sys, os, shutil, csv, hashlib, traceback, re
from datetime import datetime
from pathlib import Path
from typing import Optional, List, Tuple, Dict, Any
from dataclasses import dataclass, field
from collections import defaultdict

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QLineEdit, QPushButton, QDateEdit, QComboBox,
    QGroupBox, QScrollArea, QFrame, QFileDialog, QMessageBox,
    QSplitter, QCheckBox, QButtonGroup, QRadioButton,
    QProgressBar, QSizePolicy, QListWidget, QListWidgetItem,
    QAbstractItemView, QDialog, QTabWidget, QTableWidget,
    QTableWidgetItem, QHeaderView,
)
from PyQt6.QtCore import Qt, QDate, QTimer, pyqtSignal, QThread
from PyQt6.QtGui import QPixmap, QCloseEvent, QColor

import io
try:
    if hasattr(sys.stdout, 'buffer'):
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    if hasattr(sys.stderr, 'buffer'):
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')
except Exception:
    pass

# ── Optional dependencies ─────────────────────────────────────────────────────
try:
    import numpy as np
    from PIL import Image, ImageFilter, ImageStat, UnidentifiedImageError
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

try:
    from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
    from matplotlib.figure import Figure
    import matplotlib.pyplot as plt
    HAS_MPL = True
except ImportError:
    HAS_MPL = False

try:
    import openslide
    HAS_OPENSLIDE = True
except (ImportError, Exception):
    HAS_OPENSLIDE = False

# ═══════════════════════════════════════════════════════════════════════════════
# CONSTANTS
# ═══════════════════════════════════════════════════════════════════════════════

try:
    import openpyxl
    from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

APP_TITLE   = "MIDAS Curation System"
APP_VERSION = "v2.3"
ANATOMICAL_SITE = "MOUTH"

CATEGORIES: List[Tuple[str, str]] = [
    ("XC", "Clinical Photography"),
    ("RG", "Radiograph / OPG / CT"),
    ("GM", "General Microscopy"),
    ("SM", "Slide Microscopy"),
    ("OT", "Other"),
]

BODY_SITES: List[Tuple[str, str]] = [
    ("MANDI", "Mandible"), ("MAXIL", "Maxilla"), ("PALAT", "Palate"),
    ("BUCCA", "Buccal"),   ("LING",  "Lingual"), ("LIP",   "Lip"),
    ("TONG",  "Tongue"),   ("LN",    "Lymph Node"), ("OTHERS", "Others"),
]

MAGNIFICATIONS: List[str] = ["4x", "10x", "20x", "40x", "100x"]
CYTOLOGY_MAGS:  List[str] = ["10x", "40x"]
GM_SUBCATEGORIES: List[str] = ["HISTOPATH", "CYTOLOGY", "IHC", "SPECIAL_STAINS"]
SM_SUBCATEGORIES: List[str] = ["HISTOPATH", "IHC", "SPECIAL_STAINS", "CYTOLOGY"]
OT_SUBCATEGORIES: List[str] = ["GROSS", "GENOMIC"]

CSV_HEADERS: List[str] = [
    "UHID", "MIDAS_CODE", "VisitDate", "BodySite",
    "XC", "RG", "Gross", "Special_Stains", "IHC",
    "Cytology", "Genomic",
    "Histopath_4x", "Histopath_10x", "Histopath_20x",
    "Histopath_40x", "Histopath_100x",
    "WSI", "Curator",
]

# Standard images + WSI formats
IMAGE_EXT     = {'.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif'}
WSI_EXT       = {'.ndpi', '.svs', '.scn', '.mrxs', '.vms', '.vmu',
                 '.czi', '.lif', '.qptiff', '.btf'}
ALL_IMAGE_EXT = IMAGE_EXT | WSI_EXT

# Dataset Counter — case folder name must match  NN_NNN_P  (e.g. 02_132_P)
CASE_ID_RE = re.compile(r'^\d{2}_\d{3}_P$')

# Dataset counter column order (matches Excel output)
COUNT_COLS = [
    "XC", "RG", "Gross", "Special_Stains", "IHC",
    "Cytology", "Genomic",
    "Histopath_4x", "Histopath_10x", "Histopath_20x",
    "Histopath_40x", "Histopath_100x", "WSI",
]

# Modality codes we can detect from path segments
MODALITY_CODES = {"XC", "RG", "GM", "SM", "OT",
                  "HISTOPATH", "CYTOLOGY", "IHC", "SPECIAL_STAINS",
                  "CLINICAL", "RADIOGRAPH", "GROSS", "GENOMIC"}

# ═══════════════════════════════════════════════════════════════════════════════
# STYLESHEET
# ═══════════════════════════════════════════════════════════════════════════════

QSS = """
* { font-family:'Segoe UI','SF Pro Display','Helvetica Neue',Arial,sans-serif;
    font-size:13px; color:#E6EDF3; }
QMainWindow  { background:#0A0E17; }
QWidget#central { background:#0A0E17; }
QWidget      { background:transparent; }
QDialog      { background:#0A0E17; }

QScrollArea  { border:none; background:transparent; }
QScrollBar:vertical   { background:#161B22; width:6px; margin:0; border-radius:3px; }
QScrollBar::handle:vertical   { background:#30363D; border-radius:3px; min-height:30px; }
QScrollBar::handle:vertical:hover { background:#0ABDC6; }
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height:0; }
QScrollBar:horizontal { background:#161B22; height:6px; border-radius:3px; }
QScrollBar::handle:horizontal { background:#30363D; border-radius:3px; }
QScrollBar::handle:horizontal:hover { background:#0ABDC6; }
QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal { width:0; }

QGroupBox { background:#161B22; border:1px solid #21262D; border-radius:8px;
    margin-top:14px; padding-top:14px; padding-left:10px;
    padding-right:10px; padding-bottom:10px; }
QGroupBox::title { subcontrol-origin:margin; left:12px; padding:0 6px;
    color:#0ABDC6; font-weight:700; font-size:10px; letter-spacing:1.5px; }

QLabel        { background:transparent; color:#8B949E; font-size:12px; }
QLabel#subheading { color:#0ABDC6; font-size:10px; letter-spacing:2px; font-weight:700; }
QLabel#filename_preview { color:#0ABDC6; font-family:'Consolas','Courier New',monospace;
    font-size:11px; background:#0D1117; border:1px solid #21262D;
    border-radius:4px; padding:6px 10px; }
QLabel#counter_badge { color:#FFFFFF; font-size:12px; font-weight:800;
    background:#0ABDC6; border-radius:10px; padding:2px 10px; min-width:40px; }
QLabel#unsaved_badge { color:#FFFFFF; font-size:11px; font-weight:700;
    background:#E3A000; border-radius:8px; padding:2px 8px; }

QLineEdit { background:#0D1117; border:1px solid #30363D; border-radius:6px;
    padding:7px 10px; color:#E6EDF3; font-size:13px;
    selection-background-color:#1F6FEB; }
QLineEdit:focus { border-color:#0ABDC6; background:#0F1923; }
QLineEdit:hover { border-color:#484F58; }
QLineEdit[readOnly="true"] { color:#8B949E; }

QDateEdit { background:#0D1117; border:1px solid #30363D; border-radius:6px;
    padding:7px 10px; color:#E6EDF3; font-size:13px; }
QDateEdit:focus { border-color:#0ABDC6; }
QDateEdit::drop-down { border:none; width:22px; }
QCalendarWidget { background:#1C2128; color:#E6EDF3; }

QComboBox { background:#0D1117; border:1px solid #30363D; border-radius:6px;
    padding:7px 10px; color:#E6EDF3; font-size:13px; }
QComboBox:focus, QComboBox:on { border-color:#0ABDC6; }
QComboBox::drop-down { border:none; width:24px; }
QComboBox::down-arrow { border:4px solid transparent; border-top:5px solid #8B949E;
    width:0; height:0; margin-right:8px; }
QComboBox QAbstractItemView { background:#1C2128; border:1px solid #30363D;
    border-radius:6px; selection-background-color:#1F3A5C;
    outline:none; padding:4px; color:#E6EDF3; }

QListWidget { background:#0D1117; border:1px solid #30363D; border-radius:6px;
    color:#E6EDF3; font-size:12px; outline:none; }
QListWidget::item { padding:4px 8px; border-radius:4px; }
QListWidget::item:selected { background:#1F3A5C; color:#0ABDC6; }
QListWidget::item:hover { background:#21262D; }

QRadioButton { color:#C9D1D9; spacing:8px; font-size:13px; }
QRadioButton::indicator { width:16px; height:16px; border-radius:8px;
    border:2px solid #30363D; background:#0D1117; }
QRadioButton::indicator:checked { background:#0ABDC6; border-color:#0ABDC6; }
QRadioButton::indicator:hover  { border-color:#0ABDC6; }
QRadioButton:checked { color:#0ABDC6; font-weight:600; }

QCheckBox { color:#C9D1D9; spacing:6px; }
QCheckBox::indicator { width:16px; height:16px; border-radius:3px;
    border:2px solid #30363D; background:#0D1117; }
QCheckBox::indicator:checked { background:#0ABDC6; border-color:#0ABDC6; }

QPushButton { background:#21262D; color:#C9D1D9; border:1px solid #30363D;
    border-radius:6px; padding:8px 16px; font-size:13px; font-weight:500; }
QPushButton:hover { background:#2D333B; border-color:#484F58; color:#E6EDF3; }
QPushButton:pressed { background:#1C2128; }
QPushButton#btn_primary {
    background:qlineargradient(x1:0,y1:0,x2:0,y2:1,stop:0 #0ABDC6,stop:1 #0891B2);
    color:#FFFFFF; border:none; font-size:14px; font-weight:700;
    padding:12px 24px; border-radius:8px; }
QPushButton#btn_primary:hover {
    background:qlineargradient(x1:0,y1:0,x2:0,y2:1,stop:0 #22D3EE,stop:1 #0ABDC6);
    border:none; color:#FFFFFF; }
QPushButton#btn_primary:pressed  { background:#0891B2; }
QPushButton#btn_primary:disabled { background:#21262D; color:#484F58; border:1px solid #30363D; }

QPushButton#btn_secondary { background:transparent; color:#0ABDC6;
    border:1px solid #0ABDC6; border-radius:6px; padding:7px 14px; }
QPushButton#btn_secondary:hover { background:rgba(10,189,198,0.12); }
QPushButton#btn_secondary:disabled { color:#30363D; border-color:#30363D; }

QPushButton#btn_warning { background:transparent; color:#E3A000;
    border:1px solid #E3A000; border-radius:6px; padding:7px 14px; font-weight:600; }
QPushButton#btn_warning:hover { background:rgba(227,160,0,0.12); }
QPushButton#btn_warning:disabled { color:#30363D; border-color:#30363D; }

QPushButton#btn_danger { background:transparent; color:#F85149;
    border:1px solid #F85149; border-radius:6px; padding:6px 12px; font-size:12px; }
QPushButton#btn_danger:hover { background:rgba(248,81,73,0.1); }

QPushButton#btn_qc {
    background:qlineargradient(x1:0,y1:0,x2:0,y2:1,stop:0 #7C3AED,stop:1 #5B21B6);
    color:#FFFFFF; border:none; font-size:13px; font-weight:700;
    padding:10px 20px; border-radius:8px; }
QPushButton#btn_qc:hover {
    background:qlineargradient(x1:0,y1:0,x2:0,y2:1,stop:0 #8B5CF6,stop:1 #7C3AED); }
QPushButton#btn_qc:disabled { background:#21262D; color:#484F58; border:1px solid #30363D; }

QProgressBar { background:#161B22; border:1px solid #21262D; border-radius:4px;
    height:6px; text-align:center; color:transparent; }
QProgressBar::chunk {
    background:qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 #0ABDC6,stop:1 #00C896);
    border-radius:4px; }

QStatusBar { background:#161B22; border-top:1px solid #21262D;
    color:#8B949E; font-size:11px; }
QSplitter::handle { background:#21262D; width:1px; }

QTabWidget::pane { border:1px solid #21262D; background:#0D1117;
    border-radius:0 8px 8px 8px; }
QTabBar::tab { background:#161B22; color:#8B949E; border:1px solid #21262D;
    padding:8px 18px; font-size:12px; border-bottom:none;
    border-radius:6px 6px 0 0; margin-right:2px; }
QTabBar::tab:selected { background:#0D1117; color:#0ABDC6; font-weight:600; }
QTabBar::tab:hover { color:#C9D1D9; background:#21262D; }

QTableWidget { background:#0D1117; border:1px solid #21262D;
    gridline-color:#1C2128; color:#E6EDF3; font-size:11px;
    selection-background-color:#1F3A5C; }
QTableWidget::item { padding:4px 8px; border:none; }
QTableWidget::item:selected { background:#1F3A5C; color:#0ABDC6; }
QHeaderView::section { background:#161B22; color:#8B949E; border:none;
    border-right:1px solid #21262D; border-bottom:1px solid #21262D;
    padding:6px 10px; font-size:10px; letter-spacing:0.5px; font-weight:700; }
"""

# ═══════════════════════════════════════════════════════════════════════════════
# BACKEND — FOLDER / FILE / COUNTER / CSV / LOGGER / SESSION (unchanged)
# ═══════════════════════════════════════════════════════════════════════════════

class FolderBuilder:
    @staticmethod
    def get_path(root, midas_code, visit_date, category,
                 subcategory=None, body_site=None, magnification=None) -> Path:
        base = Path(root) / midas_code / f"VISIT_{visit_date}" / ANATOMICAL_SITE
        if category == "XC":
            return base / "XC" / "CLINICAL"
        elif category == "RG":
            return base / "RG" / "RADIOGRAPH"
        elif category in ("GM", "SM"):
            path = base / category
            if subcategory == "HISTOPATH":
                p = path / "HISTOPATH"
                if body_site:     p = p / body_site
                if magnification: p = p / magnification
                return p
            elif subcategory == "CYTOLOGY":
                p = path / "CYTOLOGY"
                if magnification: p = p / magnification
                return p
            elif subcategory in ("IHC", "SPECIAL_STAINS"):
                return path / subcategory
            return path
        elif category == "OT":
            path = base / "OT"
            if subcategory in ("GROSS", "GENOMIC"):
                return path / subcategory
            return path
        return base

    @staticmethod
    def ensure(path: Path) -> Path:
        path.mkdir(parents=True, exist_ok=True)
        return path


class FileNamer:
    @staticmethod
    def build(midas_code, visit_date, category,
              body_site=None, magnification=None, count=1, ext=".jpg") -> str:
        parts = [midas_code, f"VISIT_{visit_date}", category]
        if body_site:     parts.append(body_site)
        if magnification: parts.append(magnification)
        parts.append(f"{count:03d}")
        return "_".join(parts) + ext


class CounterManager:
    @staticmethod
    def next_count(folder: Path) -> int:
        if not folder.exists():
            return 1
        imgs = [f for f in folder.iterdir()
                if f.is_file() and f.suffix.lower() in ALL_IMAGE_EXT]
        return len(imgs) + 1


class CSVWriter:
    NUMERIC_COLS = {
        "XC", "RG", "Gross", "Special_Stains", "IHC", "Cytology", "Genomic",
        "Histopath_4x", "Histopath_10x", "Histopath_20x",
        "Histopath_40x", "Histopath_100x", "WSI",
    }

    def __init__(self, midas_folder: Path, midas_code: str):
        midas_folder.mkdir(parents=True, exist_ok=True)
        self.csv_path = midas_folder / f"{midas_code}_metadata.csv"
        if not self.csv_path.exists():
            with open(self.csv_path, 'w', newline='') as f:
                csv.writer(f).writerow(CSV_HEADERS)

    def upsert(self, new_row: dict):
        rows = self._read_all()
        key  = (str(new_row.get("MIDAS_CODE", "")),
                str(new_row.get("VisitDate",   "")))
        matched = False
        for row in rows:
            if (str(row.get("MIDAS_CODE", "")) == key[0] and
                    str(row.get("VisitDate", "")) == key[1]):
                for col in CSV_HEADERS:
                    if col in self.NUMERIC_COLS:
                        try:
                            row[col] = int(row.get(col) or 0) + int(new_row.get(col) or 0)
                        except (ValueError, TypeError):
                            pass
                    else:
                        if not row.get(col):
                            row[col] = new_row.get(col, "")
                matched = True
                break
        if not matched:
            rows.append({h: new_row.get(h, "") for h in CSV_HEADERS})
        self._write_all(rows)

    def _read_all(self) -> List[dict]:
        if not self.csv_path.exists():
            return []
        with open(self.csv_path, 'r', newline='') as f:
            return list(csv.DictReader(f))

    def _write_all(self, rows: List[dict]):
        with open(self.csv_path, 'w', newline='', encoding='utf-8') as f:
            w = csv.DictWriter(f, fieldnames=CSV_HEADERS, extrasaction='ignore')
            w.writeheader()
            w.writerows(rows)


class SessionLogger:
    def __init__(self, root: Path):
        logs_dir = root / "logs"
        logs_dir.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.log_path = logs_dir / f"curation_log_{ts}.txt"
        self._write(f"=== MIDAS Session started: {datetime.now():%Y-%m-%d %H:%M:%S} ===")

    def log(self, msg: str):
        self._write(f"[{datetime.now():%H:%M:%S}] {msg}")

    def _write(self, line: str):
        with open(self.log_path, 'a', encoding='utf-8') as f:
            f.write(line + "\n")


class SessionState:
    def __init__(self):
        self._ops: List[str] = []

    def mark_organised(self, desc: str): self._ops.append(desc)
    def mark_flushed(self):              self._ops.clear()

    @property
    def has_unsaved(self) -> bool:       return bool(self._ops)


# ═══════════════════════════════════════════════════════════════════════════════
# DATASET COUNTER — BACKEND
# ═══════════════════════════════════════════════════════════════════════════════

class DatasetCounter:
    """
    Recursively locates all folders whose name matches  NN_NNN_P  (CASE_ID_RE).
    For each case folder, counts every image by modality, derived from the
    folder path segments relative to the case root.

    MIDAS structure expected:
      <case_id>/VISIT_<date>/MOUTH/<MOD>/[SUBMOD]/[BODY_SITE]/[MAGNIFICATION]/img

    But the scanner is intentionally tolerant — it only requires that the
    case folder itself matches the pattern; the interior layout just needs to
    contain the known segment keywords (XC, RG, OT, HISTOPATH, IHC, etc.)
    somewhere in the relative path.
    """

    MAG_MAP = {"4x": "Histopath_4x", "10x": "Histopath_10x",
               "20x": "Histopath_20x", "40x": "Histopath_40x",
               "100x": "Histopath_100x"}

    @staticmethod
    def find_case_folders(root: Path) -> List[Path]:
        """Return every directory anywhere under root whose name matches NN_NNN_P."""
        found = []
        for dirpath, dirnames, _ in os.walk(str(root)):
            for d in dirnames:
                if CASE_ID_RE.match(d):
                    found.append(Path(dirpath) / d)
        # Keep only the top-most match — don't double-count nested structures
        found.sort()
        pruned: List[Path] = []
        for p in found:
            if not any(p.is_relative_to(existing) for existing in pruned):
                pruned.append(p)
        return pruned

    @classmethod
    def _classify(cls, rel_parts: Tuple[str, ...], is_wsi: bool) -> Optional[str]:
        """Map relative path segments → COUNT_COLS key or None."""
        if is_wsi:
            return "WSI"
        up = [s.upper() for s in rel_parts]
        if "XC" in up:
            return "XC"
        if "RG" in up:
            return "RG"
        if "OT" in up:
            if "GROSS" in up:
                return "Gross"
            if "GENOMIC" in up:
                return "Genomic"
            return None
        if "HISTOPATH" in up:
            # Look for a magnification folder anywhere in the path
            for part in rel_parts:
                key = cls.MAG_MAP.get(part.lower())
                if key:
                    return key
            return None   # histopath without magnification — skip
        if "IHC" in up:
            return "IHC"
        if "SPECIAL_STAINS" in up:
            return "Special_Stains"
        if "CYTOLOGY" in up:
            return "Cytology"
        return None

    @classmethod
    def count_case(cls, case_folder: Path) -> Dict:
        """Returns dict with counts, visit_dates set, and gm_body_parts counter."""
        counts         = {k: 0 for k in COUNT_COLS}
        visit_dates    = set()
        gm_body_parts: Dict[str, int] = defaultdict(int)
        _BODY_SITE_CODES = {"MANDI", "MAXIL", "PALAT", "BUCCA",
                            "LING", "LIP", "TONG", "LN", "OTHERS"}
        for img_path in case_folder.rglob("*"):
            if not img_path.is_file():
                continue
            ext = img_path.suffix.lower()
            if ext not in ALL_IMAGE_EXT:
                continue
            is_wsi = ext in WSI_EXT
            rel    = img_path.relative_to(case_folder)
            parts  = rel.parts[:-1]   # exclude filename
            # Collect visit date from path
            for p in parts:
                if p.startswith("VISIT_"):
                    visit_dates.add(p[len("VISIT_"):])
            key = cls._classify(parts, is_wsi)
            if key:
                counts[key] += 1
            # GM body part distribution — only for histopath images
            up_parts = [p.upper() for p in parts]
            if "GM" in up_parts or "SM" in up_parts:
                if "HISTOPATH" in up_parts:
                    for p in parts:
                        if p.upper() in _BODY_SITE_CODES:
                            gm_body_parts[p.upper()] += 1
                            break
        result = dict(counts)
        result["_visit_dates"]   = sorted(visit_dates)
        result["_gm_body_parts"] = dict(gm_body_parts)
        return result

    @classmethod
    def count_all(cls, root: Path,
                  progress_cb=None) -> List[Dict]:
        """
        Returns a list of dicts, one per case, sorted by case_id.
        Each dict: {case_id, case_path, institute_hint, **COUNT_COLS}
        progress_cb(done, total, case_id)
        """
        cases    = cls.find_case_folders(root)
        results  = []
        total    = len(cases)
        for i, case_folder in enumerate(cases):
            if progress_cb:
                progress_cb(i + 1, total, case_folder.name)
            data   = cls.count_case(case_folder)
            vdates = data.pop("_visit_dates", [])
            gmbp   = data.pop("_gm_body_parts", {})
            row = {
                "case_id":       case_folder.name,
                "case_path":     str(case_folder),
                "visit_dates":   ", ".join(vdates) if vdates else "",
                "gm_body_parts": gmbp,
            }
            row.update(data)
            results.append(row)
        return sorted(results, key=lambda r: r["case_id"])


class DatasetCountWorker(QThread):
    progress = pyqtSignal(int, int, str)   # done, total, case_id
    finished = pyqtSignal(list)
    error    = pyqtSignal(str)

    def __init__(self, root: Path, parent=None):
        super().__init__(parent)
        self.root = root

    def run(self):
        try:
            results = DatasetCounter.count_all(
                self.root,
                progress_cb=lambda d, t, c: self.progress.emit(d, t, c))
            self.finished.emit(results)
        except Exception:
            self.error.emit(traceback.format_exc())


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORTER
# ─────────────────────────────────────────────────────────────────────────────

class ExcelExporter:
    """Builds a styled openpyxl workbook from dataset-counter results."""

    # Palette
    HDR_FILL  = "0D1117"
    HDR_FONT  = "0ABDC6"
    ID_FILL   = "161B22"
    ID_FONT   = "0ABDC6"
    ALT_FILL  = "0F1520"
    ZERO_FONT = "30363D"
    TOT_FILL  = "00C896"
    TOT_FONT  = "0D1117"
    BORDER_C  = "21262D"

    MOD_COLORS = {
        "XC":            "0ABDC6",
        "RG":            "E3A000",
        "Gross":         "F85149",
        "Special_Stains":"7C3AED",
        "IHC":           "E879F9",
        "Cytology":      "34D399",
        "Genomic":       "1F6FEB",
        "Histopath_4x":  "00C896",
        "Histopath_10x": "00C896",
        "Histopath_20x": "00C896",
        "Histopath_40x": "00C896",
        "Histopath_100x":"00C896",
        "WSI":           "8B5CF6",
    }

    @classmethod
    def _border(cls):
        s = Side(style='thin', color=cls.BORDER_C)
        return Border(left=s, right=s, top=s, bottom=s)

    @classmethod
    def build(cls, results: List[Dict], institute: str, out_path: str):
        if not HAS_OPENPYXL:
            raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dataset Count"

        all_cols = ["Institute", "Case_ID", "Visit_Dates"] + COUNT_COLS
        n_data   = len(all_cols)

        # ── Header row ────────────────────────────────────────────────────────
        hdr_fills = {c: PatternFill("solid", fgColor=cls.MOD_COLORS.get(c, cls.HDR_FILL))
                     for c in COUNT_COLS}
        for col_i, col_name in enumerate(all_cols, start=1):
            cell = ws.cell(row=1, column=col_i, value=col_name.replace("_", " ").upper())
            if col_name in cls.MOD_COLORS:
                cell.fill = PatternFill("solid",
                                        fgColor=cls.MOD_COLORS[col_name])
                cell.font = Font(bold=True, color="0D1117", size=9,
                                 name="Consolas")
            else:
                cell.fill = PatternFill("solid", fgColor=cls.HDR_FILL)
                cell.font = Font(bold=True, color=cls.HDR_FONT, size=10,
                                 name="Segoe UI")
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            cell.border = cls._border()
        ws.row_dimensions[1].height = 36

        # ── Data rows ─────────────────────────────────────────────────────────
        for row_i, row in enumerate(results, start=2):
            is_alt = (row_i % 2 == 0)
            row_fill = PatternFill("solid", fgColor=cls.ALT_FILL if is_alt else cls.HDR_FILL)

            # Institute
            c = ws.cell(row=row_i, column=1, value=institute or "—")
            c.fill = row_fill
            c.font = Font(color="8B949E", size=10, name="Segoe UI")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = cls._border()

            # Case_ID
            c = ws.cell(row=row_i, column=2, value=row["case_id"])
            c.fill = PatternFill("solid", fgColor=cls.ID_FILL)
            c.font = Font(bold=True, color=cls.ID_FONT, size=11,
                          name="Consolas")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = cls._border()

            # Visit Dates
            c = ws.cell(row=row_i, column=3, value=row.get("visit_dates", ""))
            c.fill = row_fill
            c.font = Font(color="8B949E", size=9, name="Consolas")
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
            c.border = cls._border()

            # Count columns
            for col_i, col_name in enumerate(COUNT_COLS, start=4):
                val  = row.get(col_name, 0)
                cell = ws.cell(row=row_i, column=col_i, value=val)
                cell.fill   = row_fill
                cell.border = cls._border()
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if val == 0:
                    cell.font = Font(color=cls.ZERO_FONT, size=10, name="Consolas")
                else:
                    mod_clr = cls.MOD_COLORS.get(col_name, "E6EDF3")
                    cell.font = Font(bold=True, color=mod_clr, size=11,
                                     name="Consolas")
            ws.row_dimensions[row_i].height = 22

        # ── Totals row ────────────────────────────────────────────────────────
        tot_row = len(results) + 2
        ws.cell(row=tot_row, column=1, value="TOTAL").fill = \
            PatternFill("solid", fgColor=cls.TOT_FILL)
        ws.cell(row=tot_row, column=1).font = \
            Font(bold=True, color=cls.TOT_FONT, size=11, name="Segoe UI")
        ws.cell(row=tot_row, column=1).alignment = \
            Alignment(horizontal="center", vertical="center")
        ws.cell(row=tot_row, column=1).border = cls._border()

        ws.cell(row=tot_row, column=2, value=f"{len(results)} cases")
        ws.cell(row=tot_row, column=2).fill = \
            PatternFill("solid", fgColor=cls.TOT_FILL)
        ws.cell(row=tot_row, column=2).font = \
            Font(bold=True, color=cls.TOT_FONT, size=11, name="Consolas")
        ws.cell(row=tot_row, column=2).alignment = \
            Alignment(horizontal="center", vertical="center")
        ws.cell(row=tot_row, column=2).border = cls._border()

        ws.cell(row=tot_row, column=3, value="—").fill = PatternFill("solid", fgColor=cls.TOT_FILL)
        ws.cell(row=tot_row, column=3).font = Font(bold=True, color=cls.TOT_FONT, size=11, name="Consolas")
        ws.cell(row=tot_row, column=3).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=tot_row, column=3).border = cls._border()
        for col_i, col_name in enumerate(COUNT_COLS, start=4):
            total = sum(r.get(col_name, 0) for r in results)
            cell  = ws.cell(row=tot_row, column=col_i, value=total)
            cell.fill = PatternFill("solid", fgColor=cls.TOT_FILL)
            cell.font = Font(bold=True, color=cls.TOT_FONT, size=11,
                             name="Consolas")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = cls._border()
        ws.row_dimensions[tot_row].height = 26

        # ── Column widths ─────────────────────────────────────────────────────
        ws.column_dimensions["A"].width = 16   # Institute
        ws.column_dimensions["B"].width = 16   # Case_ID
        ws.column_dimensions["C"].width = 24   # Visit_Dates
        for col_i in range(4, len(all_cols) + 1):
            ws.column_dimensions[get_column_letter(col_i)].width = 14

        # ── Freeze header ─────────────────────────────────────────────────────
        ws.freeze_panes = "D2"

        # ── Summary sheet ─────────────────────────────────────────────────────
        ws2 = wb.create_sheet("Summary")
        ws2["A1"] = "MIDAS Dataset Summary"
        ws2["A1"].font = Font(bold=True, color="0ABDC6", size=14, name="Segoe UI")
        ws2["A2"] = f"Generated: {datetime.now():%Y-%m-%d  %H:%M:%S}"
        ws2["A2"].font = Font(color="8B949E", size=10)
        ws2["A3"] = f"Institute: {institute or '—'}"
        ws2["A3"].font = Font(color="8B949E", size=10)
        ws2["A4"] = f"Total Cases: {len(results)}"
        ws2["A4"].font = Font(bold=True, color="E6EDF3", size=11)
        ws2["A5"] = f"Root Folder: (scan root)"
        ws2["A5"].font = Font(color="8B949E", size=10)
        ws2.column_dimensions["A"].width = 40

        row_s = 7
        ws2.cell(row=row_s, column=1, value="MODALITY").font = \
            Font(bold=True, color="0ABDC6", size=10, name="Segoe UI")
        ws2.cell(row=row_s, column=2, value="TOTAL IMAGES").font = \
            Font(bold=True, color="0ABDC6", size=10, name="Segoe UI")
        ws2.cell(row=row_s, column=3, value="CASES WITH DATA").font = \
            Font(bold=True, color="0ABDC6", size=10, name="Segoe UI")

        for si, col_name in enumerate(COUNT_COLS, start=row_s + 1):
            total     = sum(r.get(col_name, 0) for r in results)
            non_zero  = sum(1 for r in results if r.get(col_name, 0) > 0)
            clr       = cls.MOD_COLORS.get(col_name, "8B949E")
            ws2.cell(row=si, column=1, value=col_name.replace("_", " ")).font = \
                Font(color=clr, size=10, name="Consolas")
            ws2.cell(row=si, column=2, value=total).font = \
                Font(bold=True, color="E6EDF3", size=11, name="Consolas")
            ws2.cell(row=si, column=3, value=non_zero).font = \
                Font(color="8B949E", size=10, name="Consolas")
        ws2.column_dimensions["B"].width = 16
        ws2.column_dimensions["C"].width = 18

        # ── GM Body Part Distribution sheet ──────────────────────────────────
        _BODY_SITE_ORDER = ["MANDI", "MAXIL", "PALAT", "BUCCA",
                            "LING", "LIP", "TONG", "LN", "OTHERS"]
        _BS_COLORS = {
            "MANDI": "0ABDC6", "MAXIL": "E3A000", "PALAT": "00C896",
            "BUCCA": "7C3AED", "LING":  "E879F9", "LIP":   "34D399",
            "TONG":  "F85149", "LN":    "1F6FEB", "OTHERS":"8B5CF6",
        }
        ws3 = wb.create_sheet("GM Body Part Distribution")
        ws3["A1"] = "GM / HISTOPATH  —  Body Part Distribution"
        ws3["A1"].font = Font(bold=True, color="00C896", size=13, name="Segoe UI")
        ws3.row_dimensions[1].height = 22
        ws3["A2"] = "Body site frequency across all cases (GM/SM Histopath images)"
        ws3["A2"].font = Font(color="8B949E", size=10, name="Segoe UI")
        ws3.row_dimensions[2].height = 18

        gm_hdr_row = 4
        gm_hdr_cols = ["Case_ID", "Visit_Dates"] + _BODY_SITE_ORDER + ["TOTAL_GM"]
        for ci, h in enumerate(gm_hdr_cols, start=1):
            c = ws3.cell(row=gm_hdr_row, column=ci, value=h.replace("_", " ").upper())
            clr = _BS_COLORS.get(h, cls.HDR_FILL)
            c.fill = PatternFill("solid", fgColor=clr if h in _BS_COLORS else cls.HDR_FILL)
            c.font = Font(bold=True,
                          color="0D1117" if h in _BS_COLORS else cls.HDR_FONT,
                          size=9, name="Consolas")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = cls._border()
        ws3.row_dimensions[gm_hdr_row].height = 32

        gm_site_totals: Dict[str, int] = {bs: 0 for bs in _BODY_SITE_ORDER}
        for ri, row in enumerate(results, start=gm_hdr_row + 1):
            gmbp = row.get("gm_body_parts", {})
            is_alt = ri % 2 == 0
            rfill  = PatternFill("solid", fgColor=cls.ALT_FILL if is_alt else cls.HDR_FILL)
            row_total_gm = sum(gmbp.get(bs, 0) for bs in _BODY_SITE_ORDER)

            # Case_ID
            c = ws3.cell(row=ri, column=1, value=row["case_id"])
            c.fill = PatternFill("solid", fgColor=cls.ID_FILL)
            c.font = Font(bold=True, color=cls.ID_FONT, size=10, name="Consolas")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = cls._border()
            # Visit dates
            c = ws3.cell(row=ri, column=2, value=row.get("visit_dates", ""))
            c.fill = rfill
            c.font = Font(color="8B949E", size=9, name="Consolas")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = cls._border()
            # Body sites
            for bsi, bs in enumerate(_BODY_SITE_ORDER, start=3):
                val = gmbp.get(bs, 0)
                gm_site_totals[bs] += val
                c = ws3.cell(row=ri, column=bsi, value=val if val else None)
                c.fill = rfill
                c.border = cls._border()
                c.alignment = Alignment(horizontal="center", vertical="center")
                bsclr = _BS_COLORS.get(bs, "E6EDF3")
                c.font = Font(bold=bool(val), color=bsclr if val else cls.ZERO_FONT,
                              size=10 if val else 9, name="Consolas")
            # Row total
            c = ws3.cell(row=ri, column=len(gm_hdr_cols), value=row_total_gm if row_total_gm else None)
            c.fill = rfill
            c.border = cls._border()
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(bold=bool(row_total_gm), color="E6EDF3" if row_total_gm else cls.ZERO_FONT,
                          size=11 if row_total_gm else 9, name="Consolas")
            ws3.row_dimensions[ri].height = 20

        # Totals row for GM sheet
        gm_tot_row = len(results) + gm_hdr_row + 1
        for ci, h in enumerate(gm_hdr_cols, start=1):
            if h == "Case_ID":
                val = "TOTAL"
            elif h == "Visit_Dates":
                val = f"{len(results)} cases"
            elif h == "TOTAL_GM":
                val = sum(gm_site_totals.values())
            elif h in _BODY_SITE_ORDER:
                val = gm_site_totals.get(h, 0)
            else:
                val = ""
            c = ws3.cell(row=gm_tot_row, column=ci, value=val)
            c.fill = PatternFill("solid", fgColor=cls.TOT_FILL)
            c.font = Font(bold=True, color=cls.TOT_FONT, size=10, name="Consolas")
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = cls._border()
        ws3.row_dimensions[gm_tot_row].height = 24

        # Column widths for GM sheet
        ws3.column_dimensions["A"].width = 16
        ws3.column_dimensions["B"].width = 22
        for ci in range(3, len(gm_hdr_cols) + 1):
            ws3.column_dimensions[get_column_letter(ci)].width = 10
        ws3.freeze_panes = "C5"

        wb.save(out_path)


# ═══════════════════════════════════════════════════════════════════════════════
# DATASET COUNT DIALOG
# ═══════════════════════════════════════════════════════════════════════════════

class DatasetCountDialog(QDialog):
    """
    Scans any root folder for MIDAS case directories (NN_NNN_P),
    counts images per modality, shows results in a table,
    and exports a styled Excel workbook.
    """

    COLS = ["Case_ID", "Visit_Dates", "XC", "RG", "Gross", "Special_Stains", "IHC",
            "Cytology", "Genomic", "Histopath_4x", "Histopath_10x",
            "Histopath_20x", "Histopath_40x", "Histopath_100x", "WSI", "TOTAL"]

    COL_COLORS = {
        "XC":            "#0ABDC6",
        "RG":            "#E3A000",
        "Gross":         "#F85149",
        "Special_Stains":"#7C3AED",
        "IHC":           "#E879F9",
        "Cytology":      "#34D399",
        "Genomic":       "#1F6FEB",
        "Histopath_4x":  "#00C896",
        "Histopath_10x": "#00C896",
        "Histopath_20x": "#00C896",
        "Histopath_40x": "#00C896",
        "Histopath_100x":"#00C896",
        "WSI":           "#8B5CF6",
        "TOTAL":         "#E6EDF3",
        "Case_ID":       "#0ABDC6",
        "Visit_Dates":   "#8B949E",
    }

    def __init__(self, initial_folder: str = "", parent=None):
        super().__init__(parent)
        self.setWindowTitle("MIDAS Dataset Counter  v2.3  —  Image Inventory")
        self.setMinimumSize(1100, 720)
        self.resize(1280, 780)
        self.setStyleSheet(QSS)
        self._results: List[Dict] = []
        self._worker: Optional[DatasetCountWorker] = None
        self._setup_ui(initial_folder)

    def _setup_ui(self, initial_folder: str):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        # ── Top bar ───────────────────────────────────────────────────────────
        top = QFrame()
        top.setStyleSheet("QFrame{background:#161B22;border-bottom:1px solid #21262D;}")
        top.setFixedHeight(72)
        tl = QHBoxLayout(top); tl.setContentsMargins(20, 10, 20, 10); tl.setSpacing(14)

        # Title
        title_v = QVBoxLayout(); title_v.setSpacing(1)
        lbl_tag = QLabel("DATASET COUNTER")
        lbl_tag.setStyleSheet("color:#00C896;font-size:10px;letter-spacing:3px;"
                              "font-weight:800;background:transparent;")
        lbl_sub = QLabel("Recursive image inventory  ·  case format  NN_NNN_P")
        lbl_sub.setStyleSheet("color:#C9D1D9;font-size:13px;font-weight:500;"
                              "background:transparent;")
        title_v.addWidget(lbl_tag); title_v.addWidget(lbl_sub)
        tl.addLayout(title_v)

        if not HAS_OPENPYXL:
            badge = QLabel("  openpyxl not installed  ")
            badge.setStyleSheet("color:white;font-size:9px;font-weight:800;"
                                f"background:{DANGER};border-radius:4px;padding:3px 6px;")
            tl.addWidget(badge)
        tl.addStretch()

        # Institute
        tl.addWidget(QLabel("Institute:"))
        self.inst_edit = QLineEdit()
        self.inst_edit.setPlaceholderText("e.g.  AIIMS Delhi")
        self.inst_edit.setFixedWidth(180)
        tl.addWidget(self.inst_edit)

        tl.addWidget(QLabel("Scan folder:"))
        self.folder_edit = QLineEdit(initial_folder)
        self.folder_edit.setPlaceholderText("Select root folder…")
        self.folder_edit.setFixedWidth(300)
        self.folder_edit.setReadOnly(True)
        tl.addWidget(self.folder_edit)

        browse = QPushButton("Browse"); browse.setFixedWidth(80)
        browse.clicked.connect(self._browse)
        tl.addWidget(browse)

        self.scan_btn = QPushButton("Scan Dataset")
        self.scan_btn.setObjectName("btn_primary")
        self.scan_btn.setFixedSize(140, 44)
        self.scan_btn.clicked.connect(self._start)
        tl.addWidget(self.scan_btn)

        self.stop_btn = QPushButton("Stop")
        self.stop_btn.setObjectName("btn_danger")
        self.stop_btn.setFixedSize(74, 44)
        self.stop_btn.setEnabled(False)
        self.stop_btn.clicked.connect(self._stop)
        tl.addWidget(self.stop_btn)

        lay.addWidget(top)

        # ── Progress strip ────────────────────────────────────────────────────
        pf = QFrame()
        pf.setStyleSheet("QFrame{background:#0D1117;border-bottom:1px solid #21262D;}")
        pf.setFixedHeight(34)
        pl = QHBoxLayout(pf); pl.setContentsMargins(20, 4, 20, 4); pl.setSpacing(12)
        self.prog_bar = QProgressBar()
        self.prog_bar.setFixedHeight(8); self.prog_bar.setTextVisible(False)
        self.prog_lbl = QLabel("Select a folder and click Scan Dataset")
        self.prog_lbl.setStyleSheet("color:#484F58;font-size:11px;background:transparent;")
        pl.addWidget(self.prog_bar, stretch=1); pl.addWidget(self.prog_lbl)
        lay.addWidget(pf)

        # ── Summary cards row ─────────────────────────────────────────────────
        card_frame = QFrame()
        card_frame.setStyleSheet("QFrame{background:#0D1117;}")
        card_frame.setFixedHeight(80)
        self.card_row = QHBoxLayout(card_frame)
        self.card_row.setContentsMargins(20, 10, 20, 10)
        self.card_row.setSpacing(12)
        self._placeholder_cards()
        lay.addWidget(card_frame)

        # ── Filter + export bar ───────────────────────────────────────────────
        fbar = QFrame()
        fbar.setStyleSheet("QFrame{background:#161B22;border-bottom:1px solid #21262D;}")
        fbar.setFixedHeight(50)
        fl = QHBoxLayout(fbar); fl.setContentsMargins(16, 8, 16, 8); fl.setSpacing(12)
        self.filter_edit = QLineEdit()
        self.filter_edit.setPlaceholderText("Filter by Case_ID…")
        self.filter_edit.setFixedWidth(220)
        self.filter_edit.textChanged.connect(self._apply_filter)
        self.hide_zeros = QCheckBox("Hide empty rows")
        self.hide_zeros.setStyleSheet("color:#C9D1D9;font-size:12px;")
        self.hide_zeros.toggled.connect(self._apply_filter)
        fl.addWidget(self.filter_edit)
        fl.addWidget(self.hide_zeros)
        fl.addStretch()
        self.export_btn = QPushButton("Export to Excel")
        self.export_btn.setObjectName("btn_primary")
        self.export_btn.setFixedSize(160, 34)
        self.export_btn.setEnabled(False)
        self.export_btn.clicked.connect(self._export)
        fl.addWidget(self.export_btn)
        lay.addWidget(fbar)

        # ── Table ─────────────────────────────────────────────────────────────
        self.table = QTableWidget(0, len(self.COLS))
        self.table.setHorizontalHeaderLabels(self.COLS)
        hdr = self.table.horizontalHeader()
        hdr.setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        hdr.setStretchLastSection(False)
        self.table.verticalHeader().hide()
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.setAlternatingRowColors(False)

        # Color header labels
        for col_i, name in enumerate(self.COLS):
            item = self.table.horizontalHeaderItem(col_i)
            if item and name in self.COL_COLORS:
                item.setForeground(QColor(self.COL_COLORS[name]))

        lay.addWidget(self.table, stretch=1)

        # ── Status strip ──────────────────────────────────────────────────────
        sb = QFrame()
        sb.setStyleSheet("QFrame{background:#161B22;border-top:1px solid #21262D;}")
        sb.setFixedHeight(30)
        sl = QHBoxLayout(sb); sl.setContentsMargins(16, 0, 16, 0)
        self.status_lbl = QLabel("Ready")
        self.status_lbl.setStyleSheet("color:#484F58;font-size:11px;background:transparent;")
        sl.addWidget(self.status_lbl)
        lay.addWidget(sb)

    def _placeholder_cards(self):
        for label, val, clr in [
            ("CASES FOUND", "—", "#E6EDF3"),
            ("TOTAL IMAGES", "—", "#E6EDF3"),
            ("MODALITIES PRESENT", "—", "#E6EDF3"),
        ]:
            self.card_row.addWidget(self._card(label, val, clr))
        self.card_row.addStretch()

    def _card(self, label, value, color="#E6EDF3") -> QFrame:
        f = QFrame()
        f.setStyleSheet("QFrame{background:#161B22;border:1px solid #21262D;border-radius:8px;}")
        fl = QVBoxLayout(f); fl.setContentsMargins(14, 8, 14, 8); fl.setSpacing(2)
        l = QLabel(label)
        l.setStyleSheet("color:#484F58;font-size:9px;letter-spacing:1px;"
                        "font-weight:700;background:transparent;")
        v = QLabel(value)
        v.setStyleSheet(f"color:{color};font-size:20px;font-weight:700;background:transparent;")
        fl.addWidget(l); fl.addWidget(v)
        return f

    def _refresh_summary_cards(self):
        while self.card_row.count():
            item = self.card_row.takeAt(0)
            if item.widget(): item.widget().deleteLater()

        n_cases   = len(self._results)
        tot_imgs  = sum(sum(r.get(c, 0) for c in COUNT_COLS)
                        for r in self._results)
        mods_pres = sum(1 for c in COUNT_COLS
                        if any(r.get(c, 0) > 0 for r in self._results))

        for label, val, clr in [
            ("CASES FOUND",       str(n_cases),  "#0ABDC6"),
            ("TOTAL IMAGES",      f"{tot_imgs:,}", "#00C896"),
            ("MODALITIES PRESENT", f"{mods_pres}/{len(COUNT_COLS)}", "#7C3AED"),
        ]:
            self.card_row.addWidget(self._card(label, val, clr))

        # Per-modality mini-cards
        for col in COUNT_COLS:
            total = sum(r.get(col, 0) for r in self._results)
            if total == 0:
                continue
            clr = self.COL_COLORS.get(col, "#8B949E")
            f = QFrame()
            f.setStyleSheet(
                f"QFrame{{background:#161B22;border:1px solid {clr}44;"
                f"border-radius:6px;border-top:2px solid {clr};}}")
            fl = QVBoxLayout(f); fl.setContentsMargins(10, 6, 10, 6); fl.setSpacing(1)
            ln = QLabel(col.replace("_", "\n").replace("Histopath\n", "HP\n"))
            ln.setStyleSheet(f"color:{clr};font-size:8px;font-weight:700;"
                             "background:transparent;")
            lv = QLabel(f"{total:,}")
            lv.setStyleSheet(f"color:{clr};font-size:13px;font-weight:700;"
                             "background:transparent;")
            fl.addWidget(ln); fl.addWidget(lv)
            self.card_row.addWidget(f)
        self.card_row.addStretch()

    # ── Scan ─────────────────────────────────────────────────────────────────

    def _browse(self):
        folder = QFileDialog.getExistingDirectory(
            self, "Select Root Folder to Scan", self.folder_edit.text() or "",
            QFileDialog.Option.ShowDirsOnly)
        if folder:
            self.folder_edit.setText(folder)

    def _start(self):
        folder = self.folder_edit.text().strip()
        if not folder or not Path(folder).is_dir():
            QMessageBox.warning(self, "No Folder",
                                "Please select a valid folder to scan."); return
        self._results.clear()
        self.table.setRowCount(0)
        self.scan_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.export_btn.setEnabled(False)
        self.prog_bar.setValue(0); self.prog_bar.setMaximum(1)
        self.status_lbl.setText("Discovering case folders…")

        self._worker = DatasetCountWorker(Path(folder))
        self._worker.progress.connect(self._on_progress)
        self._worker.finished.connect(self._on_finished)
        self._worker.error.connect(self._on_error)
        self._worker.start()

    def _stop(self):
        if self._worker:
            self._worker.terminate()
        self._scan_done()

    def _on_progress(self, done: int, total: int, case_id: str):
        self.prog_bar.setMaximum(total)
        self.prog_bar.setValue(done)
        self.prog_lbl.setText(f"Counting {done}/{total}  —  {case_id}")

    def _on_finished(self, results: List[Dict]):
        self._results = results
        self._scan_done()
        if not results:
            QMessageBox.information(
                self, "No Cases Found",
                "No folders matching the pattern  NN_NNN_P  were found.\n\n"
                "Check that your case folders are named correctly,\n"
                "e.g.  02_132_P  or  01_001_P"); return
        self._refresh_summary_cards()
        self._apply_filter()
        self.export_btn.setEnabled(True)
        self.status_lbl.setText(
            f"Scan complete  ·  {len(results)} case(s) found  ·  "
            f"ready to export")

    def _on_error(self, msg: str):
        self._scan_done()
        QMessageBox.critical(self, "Scan Error", f"Error during scan:\n\n{msg}")

    def _scan_done(self):
        self.scan_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)

    # ── Table ─────────────────────────────────────────────────────────────────

    def _apply_filter(self):
        kw         = self.filter_edit.text().strip().lower()
        hide_zeros = self.hide_zeros.isChecked()

        filtered = self._results
        if kw:
            filtered = [r for r in filtered
                        if kw in r["case_id"].lower()]
        if hide_zeros:
            filtered = [r for r in filtered
                        if sum(r.get(c, 0) for c in COUNT_COLS) > 0]

        self.table.setRowCount(len(filtered))
        for row_i, row in enumerate(filtered):
            row_total = sum(row.get(c, 0) for c in COUNT_COLS)

            # Case_ID
            ci = QTableWidgetItem(row["case_id"])
            ci.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            ci.setForeground(QColor("#0ABDC6"))
            ci.setBackground(QColor(13, 17, 23))
            self.table.setItem(row_i, 0, ci)

            # Visit_Dates
            vd = QTableWidgetItem(row.get("visit_dates", ""))
            vd.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            vd.setForeground(QColor("#8B949E"))
            vd.setBackground(QColor(13, 17, 23))
            self.table.setItem(row_i, 1, vd)

            # Count columns (now start at col 2 because Visit_Dates is col 1)
            for col_i, col_name in enumerate(COUNT_COLS, start=2):
                val  = row.get(col_name, 0)
                cell = QTableWidgetItem(str(val) if val else "—")
                cell.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                clr  = self.COL_COLORS.get(col_name, "#8B949E")
                if val == 0:
                    cell.setForeground(QColor("#30363D"))
                else:
                    cell.setForeground(QColor(clr))
                    cell.setBackground(QColor(0, 0, 0, 20))
                self.table.setItem(row_i, col_i, cell)

            # Total
            tcell = QTableWidgetItem(str(row_total))
            tcell.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            tcell.setForeground(QColor("#E6EDF3" if row_total else "#30363D"))
            if row_total > 0:
                tcell.setBackground(QColor(0, 200, 150, 18))
            self.table.setItem(row_i, len(COUNT_COLS) + 2, tcell)

        self.status_lbl.setText(
            f"Showing {len(filtered)} of {len(self._results)} case(s)")

    # ── Export ────────────────────────────────────────────────────────────────

    def _export(self):
        if not self._results:
            return
        if not HAS_OPENPYXL:
            QMessageBox.critical(
                self, "Missing Dependency",
                "openpyxl is required for Excel export.\n\n"
                "Install with:  pip install openpyxl")
            return

        institute = self.inst_edit.text().strip()
        ts        = datetime.now().strftime("%Y%m%d_%H%M%S")
        default   = f"MIDAS_Dataset_{ts}.xlsx"
        path, _   = QFileDialog.getSaveFileName(
            self, "Export Dataset Count",
            default,
            "Excel Workbook (*.xlsx)")
        if not path:
            return

        try:
            QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
            ExcelExporter.build(self._results, institute, path)
            QApplication.restoreOverrideCursor()
            QMessageBox.information(
                self, "Export Complete",
                f"Dataset count exported to:\n\n{path}\n\n"
                f"{len(self._results)} cases  ·  "
                f"{sum(sum(r.get(c,0) for c in COUNT_COLS) for r in self._results):,} images total")
        except Exception as exc:
            QApplication.restoreOverrideCursor()
            QMessageBox.critical(self, "Export Failed", str(exc))


# ═══════════════════════════════════════════════════════════════════════════════
# QC — PATH PARSER  (extract Case_ID, VISIT_Date, Modality from MIDAS structure)
# ═══════════════════════════════════════════════════════════════════════════════

def parse_midas_path(path: Path) -> Tuple[str, str, str, str]:
    """
    Walk path parts to extract (case_id, visit_date, modality, sub_modality).
    MIDAS: .../<CASE_ID>/VISIT_<DATE>/<ANAT_SITE>/<MODALITY>/<SUBMOD>/...
    sub_modality examples: RG/RADIOGRAPH, GM/HISTOPATH, GM/CYTOLOGY, XC/CLINICAL
    """
    parts = path.parts
    case_id      = ""
    visit_date   = ""
    modality     = ""
    sub_modality = ""
    _MOD    = {"XC", "RG", "GM", "SM", "OT"}
    _SUBMOD = {"CLINICAL", "RADIOGRAPH", "HISTOPATH", "CYTOLOGY",
               "IHC", "SPECIAL_STAINS", "GROSS", "GENOMIC"}
    for i, part in enumerate(parts):
        if part.startswith("VISIT_") and i > 0:
            visit_date = part[len("VISIT_"):]
            case_id    = parts[i - 1]
        pu = part.upper()
        if not modality and pu in _MOD:
            modality = pu
            if i + 1 < len(parts) and parts[i + 1].upper() in _SUBMOD:
                sub_modality = f"{pu}/{parts[i + 1].upper()}"
            else:
                sub_modality = pu
            continue
        if modality and not sub_modality and pu in _SUBMOD:
            sub_modality = f"{modality}/{pu}"
    if not modality:
        modality     = path.parent.name
        sub_modality = path.parent.name
    return case_id, visit_date, modality, sub_modality


# ═══════════════════════════════════════════════════════════════════════════════
# QC — DATA MODEL
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class ImageQCResult:
    path:       Path
    filename:   str
    folder:     str
    extension:  str
    file_size_kb: float = 0.0

    # Identifiers (parsed from MIDAS folder structure)
    case_id:    str = ""
    visit_date: str = ""
    modality:     str = ""
    sub_modality: str = ""

    # Dimensions
    width:  int = 0
    height: int = 0
    color_mode: str = ""
    is_wsi: bool = False        # NDPI / SVS / other whole-slide image

    # WSI-specific metadata
    wsi_level_count: int  = 0
    wsi_vendor:      str  = ""
    wsi_mpp:         float = 0.0   # microns per pixel (if available)

    # Quality flags
    is_corrupt:       bool = False
    is_blank:         bool = False
    is_blurry:        bool = False
    is_overexposed:   bool = False
    is_underexposed:  bool = False
    is_low_contrast:  bool = False
    is_small_file:    bool = False
    is_duplicate:     bool = False
    is_color_mismatch: bool = False

    # Pixel metrics
    blur_score:     float = 0.0
    mean_intensity: float = 0.0
    std_intensity:  float = 0.0
    dynamic_range:  int   = 0

    histogram: list = field(default_factory=list)
    error_msg: str  = ""

    @property
    def resolution_str(self) -> str:
        return f"{self.width}x{self.height}" if (self.width and self.height) else "—"

    @property
    def has_issues(self) -> bool:
        return any([self.is_corrupt, self.is_blank, self.is_blurry,
                    self.is_overexposed, self.is_underexposed,
                    self.is_low_contrast, self.is_duplicate,
                    self.is_small_file, self.is_color_mismatch])

    @property
    def issues_list(self) -> List[str]:
        flags = [
            (self.is_corrupt,        "CORRUPT"),
            (self.is_blank,          "BLANK"),
            (self.is_blurry,         "BLURRY"),
            (self.is_overexposed,    "OVEREXPOSED"),
            (self.is_underexposed,   "UNDEREXPOSED"),
            (self.is_low_contrast,   "LOW CONTRAST"),
            (self.is_duplicate,      "DUPLICATE"),
            (self.is_small_file,     "SMALL FILE"),
            (self.is_color_mismatch, "COLOR MISMATCH"),
        ]
        return [lbl for flag, lbl in flags if flag]

    @property
    def status_str(self) -> str:
        return "PASS" if not self.issues_list else " | ".join(self.issues_list)


# ═══════════════════════════════════════════════════════════════════════════════
# QC — CHECKER
# ═══════════════════════════════════════════════════════════════════════════════

class ImageQCChecker:
    BLANK_STD_THRESHOLD  = 3.0   # Only flag if ALSO nearly all-black or all-white
    BLUR_THRESHOLD       = 60.0
    OVEREXPOSED_FRACTION = 0.92
    UNDEREXPOSED_MEAN    = 12.0
    LOW_CONTRAST_RANGE   = 25
    SMALL_FILE_KB        = 2.0
    MAX_ANALYSIS_DIM     = 512

    # ── WSI / NDPI ────────────────────────────────────────────────────────────

    @classmethod
    def check_wsi(cls, path: Path) -> ImageQCResult:
        """Check an NDPI / SVS / other WSI using openslide."""
        r = ImageQCResult(
            path=path, filename=path.name,
            folder=str(path.parent.name),
            extension=path.suffix.lower(),
            file_size_kb=round(path.stat().st_size / (1024 * 1024), 2),  # MB for WSI
            is_wsi=True,
        )
        r.case_id, r.visit_date, r.modality, r.sub_modality = parse_midas_path(path)

        if r.file_size_kb < 0.01:   # < 10 KB for a WSI = definitely corrupt
            r.is_small_file = True

        if not HAS_OPENSLIDE:
            r.error_msg = (
                "openslide-python not installed — "
                "WSI dimensions/quality cannot be checked.\n"
                "Install: pip install openslide-python")
            r.color_mode = "WSI"
            return r

        try:
            slide = openslide.OpenSlide(str(path))
            r.width,  r.height = slide.dimensions   # level-0 (highest res)
            r.wsi_level_count  = slide.level_count
            r.color_mode       = "WSI"

            # Vendor / scanner metadata
            props = slide.properties
            r.wsi_vendor = props.get(openslide.PROPERTY_NAME_VENDOR, "")
            mpp_x = props.get(openslide.PROPERTY_NAME_MPP_X, "")
            if mpp_x:
                try:
                    r.wsi_mpp = float(mpp_x)
                except ValueError:
                    pass

            # Thumbnail pixel analysis (256×256 overview)
            thumb = slide.get_thumbnail((256, 256)).convert("L")
            arr   = None
            if HAS_PIL:
                import numpy as _np
                arr = _np.array(thumb, dtype=_np.float32)

            if arr is not None:
                r.mean_intensity = float(arr.mean())
                r.std_intensity  = float(arr.std())
                r.dynamic_range  = int(arr.max() - arr.min())
                hist, _ = _np.histogram(arr, bins=256, range=(0, 256))
                r.histogram = hist.tolist()

                if (r.std_intensity < cls.BLANK_STD_THRESHOLD and
                        (r.mean_intensity < 15 or r.mean_intensity > 245)):
                    r.is_blank = True
                if float((_np.array(arr) > 250).sum()) / arr.size > cls.OVEREXPOSED_FRACTION:
                    r.is_overexposed = True
                if r.mean_intensity < cls.UNDEREXPOSED_MEAN:
                    r.is_underexposed = True
                if r.dynamic_range < cls.LOW_CONTRAST_RANGE:
                    r.is_low_contrast = True

            slide.close()

        except openslide.OpenSlideError as e:
            r.is_corrupt = True
            r.error_msg  = str(e)[:150]
        except Exception as e:
            r.is_corrupt = True
            r.error_msg  = str(e)[:150]

        return r

    # ── Standard images ───────────────────────────────────────────────────────

    @classmethod
    def check(cls, path: Path, dominant_mode: str = "") -> ImageQCResult:
        """Check a standard image (JPEG / PNG / TIFF / BMP)."""
        r = ImageQCResult(
            path=path, filename=path.name,
            folder=str(path.parent.name),
            extension=path.suffix.lower(),
            file_size_kb=round(path.stat().st_size / 1024, 2),
        )
        r.case_id, r.visit_date, r.modality, r.sub_modality = parse_midas_path(path)

        if r.file_size_kb < cls.SMALL_FILE_KB:
            r.is_small_file = True

        if not HAS_PIL:
            r.error_msg = "Pillow not installed."
            return r

        try:
            img = Image.open(path)
            img.verify()
            img = Image.open(path)
            img.load()
        except Exception as e:
            r.is_corrupt = True
            r.error_msg  = str(e)[:120]
            return r

        r.width, r.height = img.size
        r.color_mode = img.mode

        if dominant_mode and img.mode != dominant_mode:
            r.is_color_mismatch = True

        try:
            thumb = img.copy()
            if max(thumb.size) > cls.MAX_ANALYSIS_DIM:
                thumb.thumbnail((cls.MAX_ANALYSIS_DIM, cls.MAX_ANALYSIS_DIM),
                                Image.Resampling.LANCZOS)
            gray = thumb.convert("L")
            arr  = np.array(gray, dtype=np.float32)

            r.mean_intensity = float(arr.mean())
            r.std_intensity  = float(arr.std())
            r.dynamic_range  = int(arr.max() - arr.min())
            hist, _ = np.histogram(arr, bins=256, range=(0, 256))
            r.histogram = hist.tolist()

            if (r.std_intensity < cls.BLANK_STD_THRESHOLD and
                    (r.mean_intensity < 15 or r.mean_intensity > 245)):
                r.is_blank = True
            if float((arr > 250).sum()) / arr.size > cls.OVEREXPOSED_FRACTION:
                r.is_overexposed = True
            if r.mean_intensity < cls.UNDEREXPOSED_MEAN:
                r.is_underexposed = True
            if r.dynamic_range < cls.LOW_CONTRAST_RANGE:
                r.is_low_contrast = True

            if not r.is_blank:
                lap_k = ImageFilter.Kernel(
                    size=3, kernel=[0, 1, 0, 1, -4, 1, 0, 1, 0],
                    scale=1, offset=128)
                lap_arr = np.array(gray.filter(lap_k), dtype=np.float32) - 128.0
                r.blur_score = float(np.var(lap_arr))
                if r.blur_score < cls.BLUR_THRESHOLD:
                    r.is_blurry = True
        except Exception as e:
            r.error_msg = str(e)[:120]

        return r

    @staticmethod
    def file_hash(path: Path) -> str:
        h = hashlib.md5()
        try:
            with open(path, "rb") as f:
                h.update(f.read(16384))
            h.update(str(path.stat().st_size).encode())
        except Exception:
            pass
        return h.hexdigest()

    @staticmethod
    def dominant_color_mode(paths: List[Path]) -> str:
        counts: Dict[str, int] = defaultdict(int)
        if not HAS_PIL:
            return ""
        for p in paths[:30]:
            try:
                with Image.open(p) as im:
                    counts[im.mode] += 1
            except Exception:
                pass
        return max(counts, key=counts.get) if counts else ""


# ═══════════════════════════════════════════════════════════════════════════════
# QC — WORKER THREAD
# ═══════════════════════════════════════════════════════════════════════════════

class QCWorker(QThread):
    progress = pyqtSignal(int, int, str)
    finished = pyqtSignal(list)
    error    = pyqtSignal(str)

    def __init__(self, root_folder: Path, parent=None):
        super().__init__(parent)
        self.root_folder = root_folder
        self._abort = False

    def abort(self):
        self._abort = True

    def run(self):
        try:
            all_files = sorted([
                p for p in self.root_folder.rglob("*")
                if p.is_file() and p.suffix.lower() in ALL_IMAGE_EXT
            ])
            if not all_files:
                self.finished.emit([]); return

            total    = len(all_files)
            results  = []
            hash_map: Dict[str, str] = {}

            std_files = [f for f in all_files if f.suffix.lower() in IMAGE_EXT]
            dom_mode  = ImageQCChecker.dominant_color_mode(std_files)

            for i, path in enumerate(all_files):
                if self._abort:
                    break
                self.progress.emit(i + 1, total, path.name)

                ext = path.suffix.lower()
                if ext in WSI_EXT:
                    r = ImageQCChecker.check_wsi(path)
                else:
                    r = ImageQCChecker.check(path, dominant_mode=dom_mode)

                # Duplicate detection
                fh = ImageQCChecker.file_hash(path)
                if fh in hash_map:
                    r.is_duplicate = True
                    r.error_msg    = f"Duplicate of: {hash_map[fh]}"
                else:
                    hash_map[fh] = path.name

                results.append(r)

            self.finished.emit(results)
        except Exception:
            self.error.emit(traceback.format_exc())


# ═══════════════════════════════════════════════════════════════════════════════
# QC UI — MATPLOTLIB CANVAS
# ═══════════════════════════════════════════════════════════════════════════════

CHART_BG  = "#0D1117"
PANEL_BG  = "#161B22"
FG_COLOR  = "#8B949E"
GRID_CLR  = "#21262D"
ACCENT    = "#0ABDC6"
ACCENT2   = "#00C896"
DANGER    = "#F85149"
WARNING   = "#E3A000"
PURPLE    = "#7C3AED"
MODALITY_PALETTE = {
    "XC": "#0ABDC6",  "RG": "#E3A000",  "GM": "#00C896",
    "SM": "#7C3AED",  "OT": "#F85149",
    "UNKNOWN": "#484F58",
}


class MplCanvas(QWidget):
    def __init__(self, width=6, height=3.5, parent=None):
        super().__init__(parent)
        if not HAS_MPL:
            lay = QVBoxLayout(self)
            lbl = QLabel("matplotlib not installed.\npip install matplotlib")
            lbl.setAlignment(Qt.AlignmentFlag.AlignCenter)
            lay.addWidget(lbl)
            self.fig = None; self.ax = None; return

        self.fig    = Figure(figsize=(width, height), facecolor=CHART_BG, tight_layout=True)
        self.canvas = FigureCanvas(self.fig)
        self.canvas.setStyleSheet("background:transparent;")
        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.addWidget(self.canvas)

    def _new_ax(self, title="", xlabel="", ylabel=""):
        self.fig.clear()
        ax = self.fig.add_subplot(111)
        ax.set_facecolor(CHART_BG)
        ax.tick_params(colors=FG_COLOR, labelsize=8)
        for sp in ax.spines.values():
            sp.set_color(GRID_CLR)
        ax.xaxis.label.set_color(FG_COLOR)
        ax.yaxis.label.set_color(FG_COLOR)
        ax.grid(True, color=GRID_CLR, linewidth=0.5, alpha=0.7, axis='y')
        if title:  ax.set_title(title,  color=ACCENT, fontsize=10, pad=8, fontweight='bold')
        if xlabel: ax.set_xlabel(xlabel, color=FG_COLOR, fontsize=8)
        if ylabel: ax.set_ylabel(ylabel, color=FG_COLOR, fontsize=8)
        return ax

    # ── Charts ────────────────────────────────────────────────────────────────

    def draw_resolution_bar(self, res_counts: Dict[str, int]):
        if not self.fig: return
        ax = self._new_ax("Overall Resolution Distribution", "Resolution", "Image Count")
        labels = list(res_counts.keys())
        values = list(res_counts.values())
        colors = [ACCENT if v == max(values) else "#1F4E6B" for v in values]
        bars = ax.bar(labels, values, color=colors, width=0.6,
                      edgecolor=GRID_CLR, linewidth=0.5)
        for bar, val in zip(bars, values):
            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.3,
                    str(val), ha='center', va='bottom', color=FG_COLOR, fontsize=7)
        ax.set_xticks(range(len(labels)))
        ax.set_xticklabels(labels, rotation=40, ha='right', fontsize=7)
        self.canvas.draw()

    def draw_modality_resolution(self, data: Dict[str, Dict[str, int]]):
        """Grouped bar chart — resolution × modality."""
        if not self.fig or not data: return
        ax = self._new_ax("Resolution Distribution per Modality",
                          "Resolution", "Image Count")
        all_res = sorted(set(r for counts in data.values() for r in counts))
        n_mod   = len(data)
        width   = 0.8 / max(n_mod, 1)
        modalities = list(data.keys())

        for m_i, mod in enumerate(modalities):
            counts  = data[mod]
            offsets = [i + (m_i - n_mod / 2 + 0.5) * width
                       for i in range(len(all_res))]
            vals    = [counts.get(r, 0) for r in all_res]
            clr     = MODALITY_PALETTE.get(mod, "#8B949E")
            bars    = ax.bar(offsets, vals, width=width * 0.9, color=clr,
                             edgecolor=CHART_BG, linewidth=0.4, label=mod)

        ax.set_xticks(range(len(all_res)))
        ax.set_xticklabels(all_res, rotation=40, ha='right', fontsize=7)
        ax.legend(facecolor=PANEL_BG, edgecolor=GRID_CLR,
                  labelcolor=FG_COLOR, fontsize=8)
        self.canvas.draw()

    def draw_format_bar(self, fmt_counts: Dict[str, int]):
        if not self.fig: return
        ax = self._new_ax("File Format Distribution", "Extension", "Count")
        labels = [k.upper() for k in fmt_counts]
        values = list(fmt_counts.values())
        palette = [ACCENT, ACCENT2, WARNING, PURPLE, DANGER, "#1F6FEB", "#E879F9", "#34D399"]
        colors  = [palette[i % len(palette)] for i in range(len(labels))]
        bars = ax.bar(labels, values, color=colors, width=0.5,
                      edgecolor=GRID_CLR, linewidth=0.5)
        for bar, val in zip(bars, values):
            ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.2,
                    str(val), ha='center', va='bottom', color='#E6EDF3',
                    fontsize=9, fontweight='bold')
        self.canvas.draw()

    def draw_histogram(self, histogram_data: List[int],
                       scan_folder: str = "", modality_filter: str = ""):
        if not self.fig: return
        title = "Aggregate Pixel Intensity Histogram"
        if modality_filter and modality_filter != "ALL":
            title += f"  [{modality_filter}]"
        ax = self._new_ax(title,
                          "Pixel Intensity (0=black → 255=white)", "Frequency")
        if histogram_data:
            x = list(range(256))
            ax.fill_between(x, histogram_data, alpha=0.35, color=ACCENT)
            ax.plot(x, histogram_data, color=ACCENT, linewidth=0.9)
            total = sum(histogram_data)
            if total > 0:
                if sum(histogram_data[:16]) > total * 0.5:
                    ax.axvspan(0, 15, alpha=0.18, color=DANGER,
                               label="Underexposed zone")
                if sum(histogram_data[240:]) > total * 0.5:
                    ax.axvspan(240, 255, alpha=0.18, color=WARNING,
                               label="Overexposed zone")
            ax.legend(fontsize=7, facecolor=PANEL_BG,
                      edgecolor=GRID_CLR, labelcolor=FG_COLOR)
        self.canvas.draw()

    def draw_quality_pie(self, counts: Dict[str, int]):
        if not self.fig: return
        ax = self._new_ax("Issue Breakdown")
        labels = list(counts.keys())
        values = list(counts.values())
        if not any(values):
            ax.text(0.5, 0.5, "No data", ha='center', va='center',
                    color=FG_COLOR, transform=ax.transAxes)
            self.canvas.draw(); return
        colors = [ACCENT2, DANGER, WARNING, "#1F6FEB", PURPLE, "#E879F9"][:len(labels)]
        wedges, texts, autotexts = ax.pie(
            values, labels=labels, colors=colors, autopct='%1.0f%%',
            startangle=90,
            wedgeprops=dict(width=0.55, edgecolor=CHART_BG),
            textprops=dict(color=FG_COLOR, fontsize=8))
        for at in autotexts:
            at.set_color('#E6EDF3'); at.set_fontsize(7)
        self.canvas.draw()

    def export_figure(self, default_name: str = "chart", parent=None):
        """Save the current figure to a file chosen by the user."""
        if not self.fig:
            return
        path, _ = QFileDialog.getSaveFileName(
            parent, "Export Chart",
            default_name,
            "PNG Image (*.png);;PDF Document (*.pdf);;SVG Image (*.svg);;TIFF Image (*.tif)")
        if not path:
            return
        try:
            self.fig.savefig(path, dpi=200, bbox_inches='tight',
                             facecolor=CHART_BG, edgecolor='none')
            QMessageBox.information(parent, "Exported", f"Chart saved to:\n{path}")
        except Exception as e:
            QMessageBox.critical(parent, "Export Failed", str(e))


# ═══════════════════════════════════════════════════════════════════════════════
# QC UI — OVERVIEW TAB
# ═══════════════════════════════════════════════════════════════════════════════

class QCSummaryTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet("background:#0D1117;")
        outer = QVBoxLayout(self)
        outer.setContentsMargins(20, 20, 20, 20)
        outer.setSpacing(14)

        self.card_row = QHBoxLayout()
        self.card_row.setSpacing(10)
        outer.addLayout(self.card_row)

        mid = QHBoxLayout(); mid.setSpacing(14)
        self.pie_canvas = MplCanvas(width=4, height=3, parent=self)
        mid.addWidget(self.pie_canvas, stretch=2)

        issue_frame = QFrame()
        issue_frame.setStyleSheet(
            "QFrame{background:#161B22;border:1px solid #21262D;border-radius:8px;}")
        ifl = QVBoxLayout(issue_frame)
        ifl.setContentsMargins(14, 14, 14, 14); ifl.setSpacing(8)
        il_t = QLabel("ISSUE SUMMARY")
        il_t.setStyleSheet("color:#0ABDC6;font-size:10px;letter-spacing:2px;"
                           "font-weight:700;background:transparent;")
        ifl.addWidget(il_t)
        sc = QScrollArea()
        sc.setWidgetResizable(True)
        sc.setStyleSheet("background:transparent;border:none;")
        self.issue_inner = QWidget()
        self.issue_inner.setStyleSheet("background:transparent;")
        self.issue_layout = QVBoxLayout(self.issue_inner)
        self.issue_layout.setSpacing(5)
        self.issue_layout.setContentsMargins(0, 0, 0, 0)
        sc.setWidget(self.issue_inner)
        ifl.addWidget(sc)
        mid.addWidget(issue_frame, stretch=3)
        outer.addLayout(mid, stretch=1)

        # Health bar
        hf = QFrame()
        hf.setStyleSheet("QFrame{background:#161B22;border:1px solid #21262D;border-radius:8px;}")
        hfl = QVBoxLayout(hf); hfl.setContentsMargins(14, 12, 14, 12); hfl.setSpacing(6)
        hl = QLabel("DATASET HEALTH SCORE")
        hl.setStyleSheet("color:#0ABDC6;font-size:10px;letter-spacing:2px;"
                         "font-weight:700;background:transparent;")
        hfl.addWidget(hl)
        hr = QHBoxLayout()
        self.health_bar = QProgressBar()
        self.health_bar.setFixedHeight(14); self.health_bar.setTextVisible(False)
        self.health_pct = QLabel("—")
        self.health_pct.setStyleSheet("color:#E6EDF3;font-size:16px;font-weight:700;"
                                       "background:transparent;min-width:55px;")
        hr.addWidget(self.health_bar); hr.addWidget(self.health_pct)
        hfl.addLayout(hr)
        outer.addWidget(hf)

    def _card(self, label, value, color="#E6EDF3") -> QFrame:
        f = QFrame()
        f.setStyleSheet("QFrame{background:#161B22;border:1px solid #21262D;border-radius:8px;}")
        fl = QVBoxLayout(f); fl.setContentsMargins(12, 10, 12, 10); fl.setSpacing(3)
        l = QLabel(label)
        l.setStyleSheet("color:#484F58;font-size:9px;letter-spacing:1px;"
                        "font-weight:700;background:transparent;")
        v = QLabel(value)
        v.setStyleSheet(f"color:{color};font-size:20px;font-weight:700;background:transparent;")
        fl.addWidget(l); fl.addWidget(v)
        return f

    def _issue_row(self, label, count, color) -> QFrame:
        f = QFrame()
        f.setStyleSheet(f"QFrame{{background:transparent;border-bottom:1px solid #21262D;}}")
        rl = QHBoxLayout(f); rl.setContentsMargins(0, 5, 0, 5)
        lbl = QLabel(label)
        lbl.setStyleSheet("color:#C9D1D9;font-size:12px;background:transparent;")
        cnt = QLabel(str(count))
        cnt.setStyleSheet(f"color:{color};font-size:13px;font-weight:700;background:transparent;")
        rl.addWidget(lbl); rl.addStretch(); rl.addWidget(cnt)
        return f

    def populate(self, results: List[ImageQCResult]):
        while self.card_row.count():
            item = self.card_row.takeAt(0)
            if item.widget(): item.widget().deleteLater()

        total    = len(results)
        wsi_n    = sum(1 for r in results if r.is_wsi)
        corrupt  = sum(1 for r in results if r.is_corrupt)
        blank    = sum(1 for r in results if r.is_blank)
        blurry   = sum(1 for r in results if r.is_blurry)
        dupes    = sum(1 for r in results if r.is_duplicate)
        issues_n = sum(1 for r in results if r.has_issues)
        clean    = total - issues_n
        pct      = int(clean / total * 100) if total else 0
        clr      = "#00C896" if pct >= 90 else ("#E3A000" if pct >= 70 else "#F85149")

        for label, val, color in [
            ("TOTAL IMAGES",  str(total),    "#E6EDF3"),
            ("STANDARD",      str(total - wsi_n), "#E6EDF3"),
            ("WSI / NDPI",    str(wsi_n),    "#7C3AED" if wsi_n else "#484F58"),
            ("CLEAN / PASS",  str(clean),    "#00C896"),
            ("FLAGGED",       str(issues_n), "#F85149" if issues_n else "#00C896"),
            ("CORRUPT",       str(corrupt),  "#F85149" if corrupt else "#484F58"),
            ("DUPLICATES",    str(dupes),    "#E3A000" if dupes else "#484F58"),
        ]:
            self.card_row.addWidget(self._card(label, val, color))

        self.health_bar.setMaximum(100)
        self.health_bar.setValue(pct)
        self.health_bar.setStyleSheet(
            f"QProgressBar{{background:#0D1117;border:1px solid #21262D;"
            f"border-radius:6px;height:14px;}}"
            f"QProgressBar::chunk{{background:{clr};border-radius:5px;}}")
        self.health_pct.setText(f"{pct}%")
        self.health_pct.setStyleSheet(
            f"color:{clr};font-size:16px;font-weight:700;background:transparent;")

        while self.issue_layout.count():
            item = self.issue_layout.takeAt(0)
            if item.widget(): item.widget().deleteLater()

        for label, count, color in [
            ("Corrupt / Unreadable",  corrupt,  "#F85149"),
            ("Blank / Empty",         blank,    "#F85149"),
            ("Blurry",                blurry,   "#E3A000"),
            ("Overexposed",           sum(1 for r in results if r.is_overexposed),  "#E3A000"),
            ("Underexposed",          sum(1 for r in results if r.is_underexposed), "#E3A000"),
            ("Low Contrast",          sum(1 for r in results if r.is_low_contrast), "#E3A000"),
            ("Duplicates",            dupes,    "#E3A000"),
            ("Small File (<2 KB)",    sum(1 for r in results if r.is_small_file),   "#E3A000"),
            ("Color Mode Mismatch",   sum(1 for r in results if r.is_color_mismatch), "#7C3AED"),
        ]:
            self.issue_layout.addWidget(self._issue_row(label, count, color))
        self.issue_layout.addStretch()

        pie = {}
        if corrupt: pie["Corrupt"] = corrupt
        if blank:   pie["Blank"]   = blank
        if blurry:  pie["Blurry"]  = blurry
        oe = sum(1 for r in results if r.is_overexposed)
        ue = sum(1 for r in results if r.is_underexposed)
        if oe:  pie["Overexp"] = oe
        if ue:  pie["Underexp"] = ue
        if dupes: pie["Duplicate"] = dupes
        self.pie_canvas.draw_quality_pie(pie if pie else {"All Clean": total})


# ═══════════════════════════════════════════════════════════════════════════════
# QC UI — RESOLUTION TAB  (overall + per-modality)
# ═══════════════════════════════════════════════════════════════════════════════

class QCResolutionTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet("background:#0D1117;")
        lay = QVBoxLayout(self)
        lay.setContentsMargins(20, 20, 20, 16)
        lay.setSpacing(12)

        # Chart switcher
        top = QHBoxLayout()
        self.view_combo = QComboBox()
        self.view_combo.setFixedWidth(220)
        self.view_combo.addItem("Overall (all modalities)")
        self.view_combo.currentIndexChanged.connect(self._on_view_change)
        export_btn = QPushButton("Export Chart")
        export_btn.setObjectName("btn_secondary")
        export_btn.setFixedWidth(120)
        export_btn.clicked.connect(lambda: self.chart.export_figure("resolution_chart", self))
        top.addWidget(QLabel("View:"))
        top.addWidget(self.view_combo)
        top.addStretch()
        top.addWidget(export_btn)
        lay.addLayout(top)

        self.chart = MplCanvas(width=8, height=3.2, parent=self)
        lay.addWidget(self.chart, stretch=2)

        # Inner tabs: overall table + per-modality table
        self.inner_tabs = QTabWidget()
        self.inner_tabs.setMaximumHeight(240)

        self.overall_table = self._make_table(
            ["Resolution", "Width px", "Height px", "Count",
             "Is Square", "Aspect Ratio"])
        self.modality_table = self._make_table(
            ["Modality", "Resolution", "Count", "Case IDs (sample)"])

        self.inner_tabs.addTab(self.overall_table,  "Overall count")
        self.inner_tabs.addTab(self.modality_table, "Per-modality breakdown")
        lay.addWidget(self.inner_tabs)

        self._results: List[ImageQCResult] = []

    def _make_table(self, headers) -> QTableWidget:
        t = QTableWidget(0, len(headers))
        t.setHorizontalHeaderLabels(headers)
        t.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        t.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        t.verticalHeader().hide()
        return t

    def _on_view_change(self, idx: int):
        if not self._results: return
        if idx == 0:
            self._draw_overall()
        else:
            mod = self.view_combo.currentText()
            self._draw_for_modality(mod)

    def populate(self, results: List[ImageQCResult]):
        self._results = results

        # Populate modality filter combo
        mods = sorted(set(r.modality for r in results if r.modality))
        self.view_combo.blockSignals(True)
        self.view_combo.clear()
        self.view_combo.addItem("Overall (all modalities)")
        for m in mods:
            self.view_combo.addItem(f"Modality: {m}")
        self.view_combo.blockSignals(False)

        self._draw_overall()
        self._fill_overall_table(results)
        self._fill_modality_table(results)

    def _draw_overall(self):
        res_counts: Dict[str, int] = defaultdict(int)
        for r in self._results:
            if r.width and r.height:
                res_counts[r.resolution_str] += 1
        if res_counts:
            top20 = dict(sorted(res_counts.items(), key=lambda x: -x[1])[:20])
            self.chart.draw_resolution_bar(top20)

    def _draw_for_modality(self, modality_label: str):
        # modality_label like "Modality: XC"
        mod = modality_label.replace("Modality: ", "").strip()
        filtered = [r for r in self._results if r.modality == mod]
        data: Dict[str, Dict[str, int]] = defaultdict(lambda: defaultdict(int))
        for r in filtered:
            if r.width and r.height:
                data[mod][r.resolution_str] += 1
        if data:
            self.chart.draw_modality_resolution(dict(data))

    def _fill_overall_table(self, results):
        res_counts: Dict[str, int] = defaultdict(int)
        for r in results:
            if r.width and r.height:
                res_counts[r.resolution_str] += 1
        rows = sorted(res_counts.items(), key=lambda x: -x[1])
        self.overall_table.setRowCount(len(rows))
        for i, (res, cnt) in enumerate(rows):
            parts = res.split("x") if "x" in res else ["0", "0"]
            w, h  = int(parts[0]) if parts[0].isdigit() else 0, \
                    int(parts[1]) if len(parts) > 1 and parts[1].isdigit() else 0
            is_sq = "Yes" if w == h else "No"
            ar    = f"{w/h:.2f}" if h else "—"
            for j, val in enumerate([res, str(w), str(h), str(cnt), is_sq, ar]):
                cell = QTableWidgetItem(val)
                cell.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if j == 3:
                    cell.setForeground(QColor(ACCENT))
                elif is_sq == "Yes" and j == 4:
                    cell.setForeground(QColor("#00C896"))
                self.overall_table.setItem(i, j, cell)

    def _fill_modality_table(self, results):
        # Build: modality → resolution → count + set of case_ids
        data: Dict[str, Dict[str, Any]] = defaultdict(lambda: defaultdict(
            lambda: {"count": 0, "cases": set()}))
        for r in results:
            if r.width and r.height:
                entry = data[r.modality or "UNKNOWN"][r.resolution_str]
                entry["count"] += 1
                if r.case_id:
                    entry["cases"].add(r.case_id)

        rows_out = []
        for mod in sorted(data.keys()):
            for res in sorted(data[mod].keys()):
                entry = data[mod][res]
                sample_cases = ", ".join(sorted(entry["cases"])[:3])
                if len(entry["cases"]) > 3:
                    sample_cases += f" (+{len(entry['cases']) - 3} more)"
                rows_out.append((mod, res, entry["count"], sample_cases))
        rows_out.sort(key=lambda x: (-x[2], x[0]))

        self.modality_table.setRowCount(len(rows_out))
        for i, (mod, res, cnt, cases) in enumerate(rows_out):
            clr = MODALITY_PALETTE.get(mod, "#8B949E")
            for j, val in enumerate([mod, res, str(cnt), cases]):
                cell = QTableWidgetItem(val)
                cell.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                if j == 0:
                    cell.setForeground(QColor(clr))
                elif j == 2:
                    cell.setForeground(QColor(ACCENT))
                self.modality_table.setItem(i, j, cell)


# ═══════════════════════════════════════════════════════════════════════════════
# QC UI — FORMAT TAB
# ═══════════════════════════════════════════════════════════════════════════════

class QCFormatTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet("background:#0D1117;")
        lay = QVBoxLayout(self)
        lay.setContentsMargins(20, 20, 20, 20)
        lay.setSpacing(14)

        top = QHBoxLayout()
        lbl = QLabel("FILE FORMAT BREAKDOWN")
        lbl.setStyleSheet("color:#0ABDC6;font-size:10px;letter-spacing:2px;"
                          "font-weight:700;background:transparent;")
        export_btn = QPushButton("Export Chart")
        export_btn.setObjectName("btn_secondary")
        export_btn.setFixedWidth(120)
        export_btn.clicked.connect(lambda: self.chart.export_figure("format_chart", self))
        top.addWidget(lbl); top.addStretch(); top.addWidget(export_btn)
        lay.addLayout(top)

        self.chart = MplCanvas(width=6, height=3.5, parent=self)
        lay.addWidget(self.chart, stretch=2)

        self.cards_row = QHBoxLayout()
        self.cards_row.setSpacing(10)
        lay.addLayout(self.cards_row)
        lay.addStretch()

    def populate(self, results: List[ImageQCResult]):
        fmt_counts: Dict[str, int] = defaultdict(int)
        for r in results:
            key = r.extension.lstrip('.').lower() or 'unknown'
            if key == 'jpeg': key = 'jpg'
            fmt_counts[key] += 1

        if not fmt_counts: return
        self.chart.draw_format_bar(fmt_counts)

        while self.cards_row.count():
            item = self.cards_row.takeAt(0)
            if item.widget(): item.widget().deleteLater()

        total  = sum(fmt_counts.values())
        colors = {"jpg": ACCENT, "png": ACCENT2, "tif": WARNING, "tiff": WARNING,
                  "bmp": PURPLE, "ndpi": "#E879F9", "svs": "#E879F9",
                  "scn": "#34D399", "unknown": DANGER}

        for ext, cnt in sorted(fmt_counts.items(), key=lambda x: -x[1]):
            pct = round(cnt / total * 100, 1) if total else 0
            clr = colors.get(ext, "#8B949E")
            is_wsi_fmt = ext in {e.lstrip('.') for e in WSI_EXT}
            card = QFrame()
            card.setStyleSheet(
                f"QFrame{{background:#161B22;border:1px solid {clr}44;"
                f"border-radius:8px;border-top:3px solid {clr};}}")
            cfl = QVBoxLayout(card)
            cfl.setContentsMargins(14, 10, 14, 10); cfl.setSpacing(2)
            el = QLabel(f".{ext.upper()}")
            el.setStyleSheet(f"color:{clr};font-size:15px;font-weight:800;background:transparent;")
            nl = QLabel(f"{cnt} files")
            nl.setStyleSheet("color:#E6EDF3;font-size:12px;font-weight:600;background:transparent;")
            pl = QLabel(f"{pct}% of dataset")
            pl.setStyleSheet("color:#8B949E;font-size:10px;background:transparent;")
            if is_wsi_fmt:
                wl = QLabel("WSI FORMAT")
                wl.setStyleSheet(f"color:{clr};font-size:9px;font-weight:700;"
                                 "background:transparent;letter-spacing:1px;")
                cfl.addWidget(wl)
            cfl.addWidget(el); cfl.addWidget(nl); cfl.addWidget(pl)
            self.cards_row.addWidget(card)
        self.cards_row.addStretch()


# ═══════════════════════════════════════════════════════════════════════════════
# QC UI — HISTOGRAM TAB  (with per-modality filter + export)
# ═══════════════════════════════════════════════════════════════════════════════

class QCHistogramTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet("background:#0D1117;")
        lay = QVBoxLayout(self)
        lay.setContentsMargins(20, 20, 20, 20)
        lay.setSpacing(12)

        info = QLabel(
            "Aggregate pixel intensity histogram across all non-corrupt images. "
            "Filter by modality to compare scanners / staining. "
            "A healthy dataset shows broad mid-tone coverage (40–220 range).")
        info.setWordWrap(True)
        info.setStyleSheet("color:#8B949E;font-size:12px;background:transparent;")
        lay.addWidget(info)

        # Controls row
        ctrl = QHBoxLayout(); ctrl.setSpacing(12)
        ctrl.addWidget(QLabel("Modality filter:"))
        self.mod_combo = QComboBox()
        self.mod_combo.setFixedWidth(200)
        self.mod_combo.addItem("ALL")
        self.mod_combo.currentIndexChanged.connect(self._on_filter_change)
        ctrl.addWidget(self.mod_combo)
        ctrl.addStretch()
        export_btn = QPushButton("Export Histogram")
        export_btn.setObjectName("btn_secondary")
        export_btn.setFixedWidth(150)
        export_btn.clicked.connect(self._export)
        ctrl.addWidget(export_btn)
        lay.addLayout(ctrl)

        self.chart = MplCanvas(width=8, height=3.8, parent=self)
        lay.addWidget(self.chart, stretch=1)

        self.stats_row = QHBoxLayout(); self.stats_row.setSpacing(10)
        lay.addLayout(self.stats_row)
        self._results: List[ImageQCResult] = []

    def _stat_card(self, label, value) -> QFrame:
        f = QFrame()
        f.setStyleSheet("QFrame{background:#161B22;border:1px solid #21262D;border-radius:6px;}")
        fl = QVBoxLayout(f); fl.setContentsMargins(12, 8, 12, 8); fl.setSpacing(2)
        l = QLabel(label)
        l.setStyleSheet("color:#484F58;font-size:10px;letter-spacing:1px;"
                        "font-weight:700;background:transparent;")
        v = QLabel(value)
        v.setStyleSheet("color:#E6EDF3;font-size:14px;font-weight:600;background:transparent;")
        fl.addWidget(l); fl.addWidget(v)
        return f

    def populate(self, results: List[ImageQCResult]):
        self._results = results
        mods = sorted(set(r.modality for r in results if r.modality))
        self.mod_combo.blockSignals(True)
        self.mod_combo.clear()
        self.mod_combo.addItem("ALL")
        for m in mods:
            self.mod_combo.addItem(m)
        self.mod_combo.blockSignals(False)
        self._redraw("ALL")

    def _on_filter_change(self):
        self._redraw(self.mod_combo.currentText())

    def _redraw(self, mod_filter: str):
        if mod_filter == "ALL":
            subset = [r for r in self._results if not r.is_corrupt and r.histogram]
        else:
            subset = [r for r in self._results
                      if not r.is_corrupt and r.histogram and r.modality == mod_filter]

        agg = [0] * 256
        for r in subset:
            for i, v in enumerate(r.histogram[:256]):
                agg[i] += v

        self.chart.draw_histogram(agg, modality_filter=mod_filter)

        # Stats
        while self.stats_row.count():
            item = self.stats_row.takeAt(0)
            if item.widget(): item.widget().deleteLater()

        means  = [r.mean_intensity for r in subset]
        stds   = [r.std_intensity  for r in subset]
        blurs  = [r.blur_score     for r in subset if r.blur_score > 0]
        wsi_n  = sum(1 for r in subset if r.is_wsi)

        stats = [
            ("IMAGES ANALYSED",    str(len(subset))),
            ("WSI IN SUBSET",      str(wsi_n)),
            ("AVG MEAN INTENSITY", f"{sum(means)/len(means):.1f}" if means else "—"),
            ("AVG STD DEV",        f"{sum(stds)/len(stds):.1f}"   if stds  else "—"),
            ("AVG BLUR SCORE",     f"{sum(blurs)/len(blurs):.1f}" if blurs else "—"),
        ]
        for label, val in stats:
            self.stats_row.addWidget(self._stat_card(label, val))
        self.stats_row.addStretch()

    def _export(self):
        mod = self.mod_combo.currentText()
        self.chart.export_figure(f"histogram_{mod}", self)


# ═══════════════════════════════════════════════════════════════════════════════
# QC UI — PER-IMAGE TABLE TAB
# ═══════════════════════════════════════════════════════════════════════════════

class QCTableTab(QWidget):
    """
    Columns (in order):
      Row# | Case_ID | VISIT_Date | Image_Name | Modality | Extension |
      Resolution | Size | Color_Mode | WSI | Blur | Mean | Issues
    """
    COLS = [
        "#", "Case_ID", "VISIT_Date", "Image_Name", "Modality", "Sub-Modality",
        "Extension", "Resolution", "Size (KB/MB)", "Mode",
        "WSI", "Blur Score", "Mean Int.", "Issues",
    ]
    # QC export CSV headers — more complete than display columns
    QC_CSV_HEADERS = [
        "Case_ID", "VISIT_Date", "Image_Name", "Modality", "Sub_Modality",
        "Extension", "Resolution", "Width_px", "Height_px",
        "File_Size", "Color_Mode",
        "Is_WSI", "WSI_Level_Count", "WSI_Vendor", "WSI_MPP_um",
        "Blur_Score", "Mean_Intensity", "Std_Dev", "Dynamic_Range",
        "Is_Corrupt", "Is_Blank", "Is_Blurry", "Is_Overexposed",
        "Is_Underexposed", "Is_Low_Contrast", "Is_Duplicate",
        "Is_Small_File", "Is_Color_Mismatch",
        "Issues", "Error_Message", "Full_Path",
    ]

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet("background:#0D1117;")
        lay = QVBoxLayout(self)
        lay.setContentsMargins(16, 16, 16, 16)
        lay.setSpacing(10)

        # Filter bar
        fr = QHBoxLayout(); fr.setSpacing(10)
        self.filter_edit = QLineEdit()
        self.filter_edit.setPlaceholderText(
            "Search by Case_ID, filename, visit date or issue…")
        self.filter_edit.setFixedWidth(320)
        self.filter_edit.textChanged.connect(self._apply_filter)

        self.mod_filter = QComboBox()
        self.mod_filter.setFixedWidth(160)
        self.mod_filter.addItem("All modalities")
        self.mod_filter.currentIndexChanged.connect(self._apply_filter)

        self.show_issues = QCheckBox("Flagged only")
        self.show_issues.setStyleSheet("color:#C9D1D9;font-size:12px;")
        self.show_issues.toggled.connect(self._apply_filter)

        self.show_wsi = QCheckBox("WSI only")
        self.show_wsi.setStyleSheet("color:#C9D1D9;font-size:12px;")
        self.show_wsi.toggled.connect(self._apply_filter)

        export_btn = QPushButton("Export QC CSV")
        export_btn.setObjectName("btn_secondary")
        export_btn.setFixedWidth(130)
        export_btn.clicked.connect(self._export_csv)

        fr.addWidget(self.filter_edit)
        fr.addWidget(self.mod_filter)
        fr.addWidget(self.show_issues)
        fr.addWidget(self.show_wsi)
        fr.addStretch()
        fr.addWidget(export_btn)
        lay.addLayout(fr)

        self.table = QTableWidget(0, len(self.COLS))
        self.table.setHorizontalHeaderLabels(self.COLS)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.verticalHeader().hide()
        lay.addWidget(self.table)

        self._all: List[ImageQCResult] = []

    def populate(self, results: List[ImageQCResult]):
        self._all = results
        # Populate modality filter
        mods = sorted(set(r.modality for r in results if r.modality))
        self.mod_filter.blockSignals(True)
        self.mod_filter.clear()
        self.mod_filter.addItem("All modalities")
        for m in mods:
            self.mod_filter.addItem(m)
        self.mod_filter.blockSignals(False)
        self._apply_filter()

    def _apply_filter(self):
        kw         = self.filter_edit.text().strip().lower()
        mod        = self.mod_filter.currentText()
        only_issues = self.show_issues.isChecked()
        only_wsi    = self.show_wsi.isChecked()

        filtered = self._all
        if only_issues: filtered = [r for r in filtered if r.has_issues]
        if only_wsi:    filtered = [r for r in filtered if r.is_wsi]
        if mod != "All modalities":
            filtered = [r for r in filtered if r.modality == mod]
        if kw:
            filtered = [r for r in filtered
                        if kw in r.filename.lower()
                        or kw in r.case_id.lower()
                        or kw in r.visit_date.lower()
                        or kw in r.status_str.lower()]

        self.table.setRowCount(len(filtered))
        for row_i, r in enumerate(filtered):
            size_str  = f"{r.file_size_kb:.1f} {'MB' if r.is_wsi else 'KB'}"
            wsi_str   = "Yes" if r.is_wsi else "—"
            vals = [
                str(row_i + 1),
                r.case_id or "—",
                r.visit_date or "—",
                r.filename,
                r.modality or "—",
                r.sub_modality or "—",
                r.extension.upper().lstrip('.'),
                r.resolution_str,
                size_str,
                r.color_mode or "—",
                wsi_str,
                f"{r.blur_score:.1f}",
                f"{r.mean_intensity:.1f}",
                r.status_str,
            ]
            if r.is_corrupt:
                bg = QColor(248, 81, 73, 30); fg_issues = QColor(DANGER)
            elif r.has_issues:
                bg = QColor(227, 160, 0, 20); fg_issues = QColor(WARNING)
            else:
                bg = QColor(0, 200, 150, 12); fg_issues = QColor(ACCENT2)

            for col_i, val in enumerate(vals):
                cell = QTableWidgetItem(val)
                cell.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                cell.setBackground(bg)
                if col_i == len(self.COLS) - 1:
                    cell.setForeground(fg_issues)
                elif col_i == 1 and r.case_id:       # Case_ID highlight
                    cell.setForeground(QColor(ACCENT))
                elif col_i == 9 and r.is_wsi:         # WSI flag
                    cell.setForeground(QColor(PURPLE))
                self.table.setItem(row_i, col_i, cell)

    def _export_csv(self):
        if not self._all:
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Export QC Report",
            f"qc_report_{datetime.now():%Y%m%d_%H%M%S}.csv",
            "CSV Files (*.csv)")
        if not path:
            return
        try:
            with open(path, 'w', newline='', encoding='utf-8') as f:
                w = csv.DictWriter(f, fieldnames=self.QC_CSV_HEADERS)
                w.writeheader()
                for r in self._all:
                    size_str = f"{r.file_size_kb:.2f} {'MB' if r.is_wsi else 'KB'}"
                    w.writerow({
                        "Case_ID":          r.case_id or "",
                        "VISIT_Date":       r.visit_date or "",
                        "Image_Name":       r.filename,
                        "Modality":         r.modality or "",
                        "Sub_Modality":     r.sub_modality or "",
                        "Extension":        r.extension.upper().lstrip('.'),
                        "Resolution":       r.resolution_str,
                        "Width_px":         r.width,
                        "Height_px":        r.height,
                        "File_Size":        size_str,
                        "Color_Mode":       r.color_mode,
                        "Is_WSI":           r.is_wsi,
                        "WSI_Level_Count":  r.wsi_level_count,
                        "WSI_Vendor":       r.wsi_vendor,
                        "WSI_MPP_um":       r.wsi_mpp,
                        "Blur_Score":       round(r.blur_score, 2),
                        "Mean_Intensity":   round(r.mean_intensity, 2),
                        "Std_Dev":          round(r.std_intensity, 2),
                        "Dynamic_Range":    r.dynamic_range,
                        "Is_Corrupt":       r.is_corrupt,
                        "Is_Blank":         r.is_blank,
                        "Is_Blurry":        r.is_blurry,
                        "Is_Overexposed":   r.is_overexposed,
                        "Is_Underexposed":  r.is_underexposed,
                        "Is_Low_Contrast":  r.is_low_contrast,
                        "Is_Duplicate":     r.is_duplicate,
                        "Is_Small_File":    r.is_small_file,
                        "Is_Color_Mismatch": r.is_color_mismatch,
                        "Issues":           r.status_str,
                        "Error_Message":    r.error_msg,
                        "Full_Path":        str(r.path),
                    })
            QMessageBox.information(self, "Exported",
                                    f"QC report saved to:\n{path}")
        except Exception as e:
            QMessageBox.critical(self, "Export Failed", str(e))


# ═══════════════════════════════════════════════════════════════════════════════
# QC UI — WSI / NDPI TAB
# ═══════════════════════════════════════════════════════════════════════════════

class QCWSITab(QWidget):
    """Dedicated tab for whole-slide image summary."""
    COLS = [
        "Case_ID", "VISIT_Date", "Filename", "Modality",
        "Format", "Dimensions", "Levels", "Vendor",
        "MPP (µm/px)", "File Size (MB)", "Issues",
    ]

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet("background:#0D1117;")
        lay = QVBoxLayout(self)
        lay.setContentsMargins(20, 20, 20, 20)
        lay.setSpacing(14)

        # Status banner
        self.status_frame = QFrame()
        self.status_frame.setStyleSheet(
            "QFrame{background:#161B22;border:1px solid #21262D;border-radius:8px;}")
        sfl = QHBoxLayout(self.status_frame)
        sfl.setContentsMargins(14, 10, 14, 10)
        self.openslide_lbl = QLabel(
            f"OpenSlide: {'AVAILABLE' if HAS_OPENSLIDE else 'NOT INSTALLED'}")
        clr = ACCENT2 if HAS_OPENSLIDE else DANGER
        self.openslide_lbl.setStyleSheet(
            f"color:{clr};font-size:13px;font-weight:700;background:transparent;")
        install_lbl = QLabel(
            "" if HAS_OPENSLIDE
            else "   Install: pip install openslide-python  (also needs OpenSlide C library)")
        install_lbl.setStyleSheet("color:#8B949E;font-size:11px;background:transparent;")
        self.wsi_count_lbl = QLabel("WSI files found: 0")
        self.wsi_count_lbl.setStyleSheet(
            f"color:{PURPLE};font-size:13px;font-weight:700;background:transparent;")
        sfl.addWidget(self.openslide_lbl)
        sfl.addWidget(install_lbl)
        sfl.addStretch()
        sfl.addWidget(self.wsi_count_lbl)
        lay.addWidget(self.status_frame)

        # Summary cards
        self.card_row = QHBoxLayout(); self.card_row.setSpacing(10)
        lay.addLayout(self.card_row)

        # Table
        self.table = QTableWidget(0, len(self.COLS))
        self.table.setHorizontalHeaderLabels(self.COLS)
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.ResizeToContents)
        self.table.horizontalHeader().setStretchLastSection(True)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.verticalHeader().hide()
        lay.addWidget(self.table)

    def _card(self, label, value, color=PURPLE) -> QFrame:
        f = QFrame()
        f.setStyleSheet(
            f"QFrame{{background:#161B22;border:1px solid {color}44;"
            f"border-radius:8px;border-left:3px solid {color};}}")
        fl = QVBoxLayout(f); fl.setContentsMargins(12, 10, 12, 10); fl.setSpacing(3)
        l = QLabel(label)
        l.setStyleSheet("color:#484F58;font-size:9px;letter-spacing:1px;"
                        "font-weight:700;background:transparent;")
        v = QLabel(value)
        v.setStyleSheet(f"color:{color};font-size:20px;font-weight:700;background:transparent;")
        fl.addWidget(l); fl.addWidget(v)
        return f

    def populate(self, results: List[ImageQCResult]):
        wsi_results = [r for r in results if r.is_wsi]
        self.wsi_count_lbl.setText(f"WSI files found: {len(wsi_results)}")

        while self.card_row.count():
            item = self.card_row.takeAt(0)
            if item.widget(): item.widget().deleteLater()

        corrupt_wsi = sum(1 for r in wsi_results if r.is_corrupt)
        formats     = len(set(r.extension for r in wsi_results))
        vendors     = set(r.wsi_vendor for r in wsi_results if r.wsi_vendor)
        ndpi_n      = sum(1 for r in wsi_results if r.extension == '.ndpi')

        for label, val, clr in [
            ("WSI TOTAL",   str(len(wsi_results)), PURPLE),
            ("NDPI FILES",  str(ndpi_n),           "#E879F9"),
            ("CORRUPT",     str(corrupt_wsi),       DANGER if corrupt_wsi else "#484F58"),
            ("FORMATS",     str(formats),           ACCENT),
            ("VENDORS",     str(len(vendors)),      ACCENT2),
        ]:
            self.card_row.addWidget(self._card(label, val, clr))
        self.card_row.addStretch()

        self.table.setRowCount(len(wsi_results))
        for i, r in enumerate(wsi_results):
            dim_str  = r.resolution_str
            mpp_str  = f"{r.wsi_mpp:.3f}" if r.wsi_mpp else "—"
            size_str = f"{r.file_size_kb:.1f}"
            vals = [
                r.case_id or "—", r.visit_date or "—", r.filename,
                r.modality or "—",
                r.extension.upper().lstrip('.'),
                dim_str,
                str(r.wsi_level_count) if r.wsi_level_count else "—",
                r.wsi_vendor or "—",
                mpp_str,
                size_str,
                r.status_str,
            ]
            bg = QColor(248, 81, 73, 30) if r.is_corrupt else QColor(124, 58, 237, 15)
            for j, val in enumerate(vals):
                cell = QTableWidgetItem(val)
                cell.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                cell.setBackground(bg)
                if j == 0:
                    cell.setForeground(QColor(ACCENT))
                elif j == len(self.COLS) - 1:
                    cell.setForeground(QColor(DANGER if r.is_corrupt else ACCENT2))
                self.table.setItem(i, j, cell)


# ═══════════════════════════════════════════════════════════════════════════════
# QC UI — MAIN DIALOG
# ═══════════════════════════════════════════════════════════════════════════════

class QCReportDialog(QDialog):
    def __init__(self, initial_folder: str = "", parent=None):
        super().__init__(parent)
        self.setWindowTitle("MIDAS QC Scanner  v2.3  —  Image Quality Control")
        self.setMinimumSize(1200, 800)
        self.resize(1380, 880)
        self.setStyleSheet(QSS)
        self._worker:  Optional[QCWorker] = None
        self._results: List[ImageQCResult] = []
        self._setup_ui(initial_folder)

    def _setup_ui(self, initial_folder: str):
        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        # Top bar
        top = QFrame()
        top.setStyleSheet("QFrame{background:#161B22;border-bottom:1px solid #21262D;}")
        top.setFixedHeight(68)
        tl = QHBoxLayout(top); tl.setContentsMargins(20, 10, 20, 10); tl.setSpacing(12)

        title_block = QVBoxLayout(); title_block.setSpacing(1)
        t1 = QLabel("QC SCANNER")
        t1.setStyleSheet("color:#7C3AED;font-size:11px;letter-spacing:3px;"
                         "font-weight:800;background:transparent;")
        t2 = QLabel("Medical Image Quality Control  ·  v2.3")
        t2.setStyleSheet("color:#C9D1D9;font-size:13px;font-weight:500;background:transparent;")
        title_block.addWidget(t1); title_block.addWidget(t2)
        tl.addLayout(title_block)

        # NDPI badge
        if HAS_OPENSLIDE:
            ndpi_badge = QLabel("  NDPI/WSI READY  ")
            ndpi_badge.setStyleSheet(
                "color:white;font-size:9px;font-weight:800;letter-spacing:1px;"
                f"background:{PURPLE};border-radius:4px;padding:3px 6px;")
            tl.addWidget(ndpi_badge)

        tl.addStretch()
        tl.addWidget(QLabel("Scan folder:"))
        self.folder_edit = QLineEdit(initial_folder)
        self.folder_edit.setPlaceholderText("Select root folder to scan…")
        self.folder_edit.setFixedWidth(360)
        self.folder_edit.setReadOnly(True)
        tl.addWidget(self.folder_edit)
        browse = QPushButton("Browse"); browse.setFixedWidth(80)
        browse.clicked.connect(self._browse)
        tl.addWidget(browse)
        self.scan_btn = QPushButton("Start Scan")
        self.scan_btn.setObjectName("btn_qc")
        self.scan_btn.setFixedSize(130, 44)
        self.scan_btn.clicked.connect(self._start)
        tl.addWidget(self.scan_btn)
        self.stop_btn = QPushButton("Stop")
        self.stop_btn.setObjectName("btn_danger")
        self.stop_btn.setFixedSize(74, 44)
        self.stop_btn.setEnabled(False)
        self.stop_btn.clicked.connect(self._stop)
        tl.addWidget(self.stop_btn)
        lay.addWidget(top)

        # Progress bar
        pf = QFrame()
        pf.setStyleSheet("QFrame{background:#0D1117;border-bottom:1px solid #21262D;}")
        pf.setFixedHeight(36)
        pl = QHBoxLayout(pf); pl.setContentsMargins(20, 5, 20, 5); pl.setSpacing(12)
        self.progress_bar = QProgressBar()
        self.progress_bar.setFixedHeight(8); self.progress_bar.setTextVisible(False)
        self.progress_lbl = QLabel("Select a folder and click Start Scan")
        self.progress_lbl.setStyleSheet("color:#484F58;font-size:11px;background:transparent;")
        pl.addWidget(self.progress_bar, stretch=1)
        pl.addWidget(self.progress_lbl)
        lay.addWidget(pf)

        # Tabs
        self.tabs = QTabWidget()
        self.summary_tab    = QCSummaryTab()
        self.resolution_tab = QCResolutionTab()
        self.format_tab     = QCFormatTab()
        self.histogram_tab  = QCHistogramTab()
        self.wsi_tab        = QCWSITab()
        self.table_tab      = QCTableTab()

        self.tabs.addTab(self.summary_tab,    "  Overview  ")
        self.tabs.addTab(self.resolution_tab, "  Resolution  ")
        self.tabs.addTab(self.format_tab,     "  Formats  ")
        self.tabs.addTab(self.histogram_tab,  "  Histogram  ")
        self.tabs.addTab(self.wsi_tab,        "  WSI / NDPI  ")
        self.tabs.addTab(self.table_tab,      "  Per-Image Report  ")
        lay.addWidget(self.tabs, stretch=1)

    def _browse(self):
        folder = QFileDialog.getExistingDirectory(
            self, "Select Folder to Scan", self.folder_edit.text() or "",
            QFileDialog.Option.ShowDirsOnly)
        if folder:
            self.folder_edit.setText(folder)

    def _start(self):
        folder = self.folder_edit.text().strip()
        if not folder or not Path(folder).is_dir():
            QMessageBox.warning(self, "No Folder",
                                "Please select a valid folder to scan."); return
        if not HAS_PIL:
            QMessageBox.critical(self, "Missing Dependency",
                                 "Pillow + numpy required:\n  pip install Pillow numpy matplotlib")
            return
        self._results.clear()
        self.scan_btn.setEnabled(False)
        self.stop_btn.setEnabled(True)
        self.progress_bar.setValue(0); self.progress_bar.setMaximum(1)

        self._worker = QCWorker(Path(folder))
        self._worker.progress.connect(self._on_progress)
        self._worker.finished.connect(self._on_finished)
        self._worker.error.connect(self._on_error)
        self._worker.start()

    def _stop(self):
        if self._worker: self._worker.abort()
        self._done()

    def _on_progress(self, done, total, name):
        self.progress_bar.setMaximum(total)
        self.progress_bar.setValue(done)
        self.progress_lbl.setText(f"Scanning {done}/{total}  —  {name}")

    def _on_finished(self, results):
        self._results = results
        self._done()
        self._populate_all()

    def _on_error(self, msg):
        self._done()
        QMessageBox.critical(self, "Scan Error", f"Error:\n\n{msg}")

    def _done(self):
        self.scan_btn.setEnabled(True)
        self.stop_btn.setEnabled(False)
        self.progress_lbl.setText(
            f"Scan complete  —  {len(self._results)} images processed")

    def _populate_all(self):
        if not self._results:
            QMessageBox.information(self, "No Images",
                                    "No images found in the selected folder.")
            return
        QApplication.setOverrideCursor(Qt.CursorShape.WaitCursor)
        try:
            self.summary_tab.populate(self._results)
            self.resolution_tab.populate(self._results)
            self.format_tab.populate(self._results)
            self.histogram_tab.populate(self._results)
            self.wsi_tab.populate(self._results)
            self.table_tab.populate(self._results)
            self.tabs.setCurrentIndex(0)
        finally:
            QApplication.restoreOverrideCursor()


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN APP — THUMBNAIL CARD
# ═══════════════════════════════════════════════════════════════════════════════

class ThumbnailCard(QFrame):
    selection_changed = pyqtSignal()
    THUMB = 138; CARD = 158

    def __init__(self, image_path: Path, parent=None):
        super().__init__(parent)
        self.image_path = image_path
        self._setup_ui(); self._load_image(); self._update_style()

    def _setup_ui(self):
        self.setFixedSize(self.CARD, self.CARD + 30)
        self.setCursor(Qt.CursorShape.PointingHandCursor)
        lay = QVBoxLayout(self)
        lay.setContentsMargins(6, 6, 6, 4); lay.setSpacing(4)
        self.img_label = QLabel()
        self.img_label.setFixedSize(self.THUMB, self.THUMB)
        self.img_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        lay.addWidget(self.img_label)
        name = self.image_path.name
        if len(name) > 22: name = name[:19] + "..."
        self.name_label = QLabel(name)
        self.name_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.name_label.setStyleSheet("color:#8B949E;font-size:10px;background:transparent;")
        lay.addWidget(self.name_label)
        self.checkbox = QCheckBox(self)
        self.checkbox.setGeometry(self.CARD - 28, 8, 20, 20)
        self.checkbox.setStyleSheet("""
            QCheckBox::indicator{width:18px;height:18px;border-radius:4px;
                border:2px solid rgba(255,255,255,0.3);background:rgba(0,0,0,0.5);}
            QCheckBox::indicator:checked{background:#0ABDC6;border-color:#0ABDC6;}
            QCheckBox{background:transparent;}""")
        self.checkbox.stateChanged.connect(self._on_check)
        self._done_badge = QLabel("ORGANISED", self)
        self._done_badge.setGeometry(0, 0, self.CARD, 22)
        self._done_badge.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self._done_badge.setStyleSheet(
            "background:rgba(0,200,150,0.88);color:white;"
            "font-size:10px;font-weight:800;")
        self._done_badge.hide()

    def mark_organised(self):
        self._done_badge.show()
        self.checkbox.setChecked(False); self.checkbox.setEnabled(False)
        self._update_style(done=True)

    def _load_image(self):
        pix = QPixmap(str(self.image_path))
        if not pix.isNull():
            pix = pix.scaled(self.THUMB, self.THUMB,
                             Qt.AspectRatioMode.KeepAspectRatio,
                             Qt.TransformationMode.SmoothTransformation)
            self.img_label.setPixmap(pix)
        else:
            self.img_label.setText("No Preview")
            self.img_label.setStyleSheet("color:#F85149;font-size:11px;background:transparent;")

    def _update_style(self, done=False):
        if done:
            self.setStyleSheet("QFrame{background:#0D2518;border:2px solid #00C896;border-radius:8px;}")
        elif self.is_selected:
            self.setStyleSheet("QFrame{background:#0D2535;border:2px solid #0ABDC6;border-radius:8px;}")
        else:
            self.setStyleSheet("QFrame{background:#1C2128;border:1px solid #30363D;border-radius:8px;}"
                               "QFrame:hover{border-color:#484F58;background:#21262D;}")

    def _on_check(self):
        self._update_style(); self.selection_changed.emit()

    def mousePressEvent(self, event):
        if event.button() == Qt.MouseButton.LeftButton and self.checkbox.isEnabled():
            self.checkbox.setChecked(not self.checkbox.isChecked())

    @property
    def is_selected(self) -> bool: return self.checkbox.isChecked()


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN APP — IMAGE PREVIEW PANEL
# ═══════════════════════════════════════════════════════════════════════════════

class ImagePreviewPanel(QWidget):
    selection_count_changed = pyqtSignal(int)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.source_folder: Optional[Path] = None
        self.thumbnails: List[ThumbnailCard] = []
        self._setup_ui()

    def _setup_ui(self):
        from PyQt6.QtWidgets import QGridLayout
        lay = QVBoxLayout(self); lay.setContentsMargins(0,0,0,0); lay.setSpacing(10)

        hdr = QHBoxLayout()
        title = QLabel("IMAGE PREVIEW"); title.setObjectName("subheading")
        hdr.addWidget(title); hdr.addStretch()
        clbl = QLabel("SELECTED")
        clbl.setStyleSheet("color:#484F58;font-size:10px;font-weight:700;"
                           "letter-spacing:1px;background:transparent;")
        self.sel_badge = QLabel("0"); self.sel_badge.setObjectName("counter_badge")
        self.sel_badge.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.sel_badge.setMinimumWidth(36)
        hdr.addWidget(clbl); hdr.addSpacing(6); hdr.addWidget(self.sel_badge)
        lay.addLayout(hdr)

        sf = QFrame()
        sf.setStyleSheet("QFrame{background:#161B22;border:1px solid #21262D;border-radius:8px;}")
        sr = QHBoxLayout(sf); sr.setContentsMargins(12,8,12,8); sr.setSpacing(10)
        self.source_label = QLabel("No source selected — browse folder")
        self.source_label.setStyleSheet("color:#484F58;font-size:11px;background:transparent;")
        sr.addWidget(self.source_label, stretch=1)
        self.browse_btn = QPushButton("Browse"); self.browse_btn.setObjectName("btn_secondary")
        self.browse_btn.setFixedWidth(120); self.browse_btn.clicked.connect(self.browse_source)
        sr.addWidget(self.browse_btn)
        self.refresh_btn = QPushButton("Refresh"); self.refresh_btn.setFixedWidth(90)
        self.refresh_btn.clicked.connect(self.refresh_images)
        sr.addWidget(self.refresh_btn)
        lay.addWidget(sf)

        acts = QHBoxLayout()
        self.select_all_btn = QPushButton("Select All"); self.select_all_btn.setFixedWidth(100)
        self.select_all_btn.clicked.connect(self.select_all)
        self.deselect_btn   = QPushButton("Deselect All"); self.deselect_btn.setFixedWidth(100)
        self.deselect_btn.clicked.connect(self.deselect_all)
        acts.addWidget(self.select_all_btn); acts.addWidget(self.deselect_btn)
        acts.addStretch()
        self.total_label = QLabel("Total: 0")
        self.total_label.setStyleSheet("color:#484F58;font-size:11px;background:transparent;")
        acts.addWidget(self.total_label)
        lay.addLayout(acts)

        self.scroll = QScrollArea(); self.scroll.setWidgetResizable(True)
        self.scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.grid_widget = QWidget(); self.grid_widget.setStyleSheet("background:transparent;")
        self.grid_layout = QGridLayout(self.grid_widget)
        self.grid_layout.setSpacing(10); self.grid_layout.setContentsMargins(8,8,8,8)
        self.grid_layout.setAlignment(Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignLeft)
        self.scroll.setWidget(self.grid_widget)

        self.empty_label = QLabel("Select a source folder to load images")
        self.empty_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.empty_label.setStyleSheet("color:#484F58;font-size:14px;background:transparent;")
        lay.addWidget(self.empty_label, stretch=1)
        self.scroll.hide()
        lay.addWidget(self.scroll, stretch=1)

    def browse_source(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Source Folder", "",
                                                  QFileDialog.Option.ShowDirsOnly)
        if folder:
            self.source_folder = Path(folder)
            self.source_label.setText(str(self.source_folder))
            self.source_label.setStyleSheet("color:#C9D1D9;font-size:11px;background:transparent;")
            self.refresh_images()

    def refresh_images(self):
        if not self.source_folder: return
        self._clear_grid()
        files = sorted([f for f in self.source_folder.iterdir()
                        if f.is_file() and f.suffix.lower() in ALL_IMAGE_EXT])
        self.total_label.setText(f"Total: {len(files)}")
        if not files:
            self.empty_label.setText(f"No images in:\n{self.source_folder}")
            self.empty_label.show(); self.scroll.hide(); self._set_badge(0); return
        self.empty_label.hide(); self.scroll.show()
        cols = 4
        for i, p in enumerate(files):
            card = ThumbnailCard(p)
            card.selection_changed.connect(self._update_count)
            self.thumbnails.append(card)
            self.grid_layout.addWidget(card, i // cols, i % cols)
        self._update_count()

    def select_all(self):
        for t in self.thumbnails:
            if t.checkbox.isEnabled(): t.checkbox.setChecked(True)

    def deselect_all(self):
        for t in self.thumbnails: t.checkbox.setChecked(False)

    def get_selected_paths(self): return [t.image_path for t in self.thumbnails if t.is_selected]
    def mark_selected_organised(self):
        for t in self.thumbnails:
            if t.is_selected: t.mark_organised()

    def _clear_grid(self):
        self.thumbnails.clear()
        while self.grid_layout.count():
            item = self.grid_layout.takeAt(0)
            if item.widget(): item.widget().deleteLater()

    def _update_count(self):
        n = sum(1 for t in self.thumbnails if t.is_selected)
        self._set_badge(n); self.selection_count_changed.emit(n)

    def _set_badge(self, n):
        self.sel_badge.setText(str(n))
        clr = "#0ABDC6" if n > 0 else "#21262D"
        tc  = "#FFFFFF" if n > 0 else "#8B949E"
        self.sel_badge.setStyleSheet(
            f"color:{tc};font-size:12px;font-weight:800;background:{clr};"
            "border-radius:10px;padding:2px 10px;min-width:40px;")

    @property
    def current_source(self) -> Optional[Path]: return self.source_folder


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN APP — BODY SITE SELECTOR
# ═══════════════════════════════════════════════════════════════════════════════

class BodySiteSelector(QWidget):
    changed = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)
        lay = QVBoxLayout(self); lay.setContentsMargins(0,0,0,0); lay.setSpacing(6)
        self.multi_check = QCheckBox("Multiple body sites for this visit")
        self.multi_check.setStyleSheet("color:#C9D1D9;font-size:12px;")
        self.multi_check.toggled.connect(self._toggle_mode)
        lay.addWidget(self.multi_check)

        self.single_combo = QComboBox()
        for code, name in BODY_SITES:
            self.single_combo.addItem(f"{code}  —  {name}", userData=code)
        self.single_combo.currentIndexChanged.connect(self.changed)
        lay.addWidget(self.single_combo)

        self.multi_list = QListWidget()
        self.multi_list.setSelectionMode(QAbstractItemView.SelectionMode.MultiSelection)
        self.multi_list.setFixedHeight(140)
        for code, name in BODY_SITES:
            item = QListWidgetItem(f"{code}  —  {name}")
            item.setData(Qt.ItemDataRole.UserRole, code)
            self.multi_list.addItem(item)
        self.multi_list.itemSelectionChanged.connect(self.changed)
        lay.addWidget(self.multi_list); self.multi_list.hide()

        self.hint = QLabel("Hold Ctrl / Cmd to select multiple sites")
        self.hint.setStyleSheet("color:#484F58;font-size:10px;background:transparent;")
        lay.addWidget(self.hint); self.hint.hide()

    def _toggle_mode(self, checked):
        self.single_combo.setVisible(not checked)
        self.multi_list.setVisible(checked)
        self.hint.setVisible(checked)
        self.changed.emit()

    def get_sites(self):
        if self.multi_check.isChecked():
            return [i.data(Qt.ItemDataRole.UserRole) for i in self.multi_list.selectedItems()]
        code = self.single_combo.currentData()
        return [code] if code else []


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN APP — FORM PANEL
# ═══════════════════════════════════════════════════════════════════════════════

class FormPanel(QWidget):
    params_changed = pyqtSignal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFixedWidth(400)
        self._setup_ui()

    def _setup_ui(self):
        outer = QVBoxLayout(self); outer.setContentsMargins(0,0,0,0)
        scroll = QScrollArea(); scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        outer.addWidget(scroll)
        content = QWidget(); content.setStyleSheet("background:transparent;")
        scroll.setWidget(content)
        lay = QVBoxLayout(content); lay.setContentsMargins(12,12,14,12); lay.setSpacing(12)

        hdr = QFrame()
        hdr.setStyleSheet("QFrame{background:qlineargradient(x1:0,y1:0,x2:1,y2:1,"
                          "stop:0 #0ABDC6,stop:1 #00C896);border-radius:10px;}")
        hfl = QVBoxLayout(hdr); hfl.setContentsMargins(16,14,16,14); hfl.setSpacing(2)
        t1 = QLabel("MIDAS")
        t1.setStyleSheet("color:white;font-size:24px;font-weight:900;"
                         "letter-spacing:5px;background:transparent;")
        t2 = QLabel("Image Curation System  ·  v2.3")
        t2.setStyleSheet("color:rgba(255,255,255,0.75);font-size:11px;"
                         "letter-spacing:1px;background:transparent;")
        hfl.addWidget(t1); hfl.addWidget(t2)
        lay.addWidget(hdr)

        g1 = QGroupBox("STEP 1  ·  ROOT STORAGE")
        v1 = QVBoxLayout(g1); v1.setSpacing(6)
        row = QHBoxLayout()
        self.root_edit = QLineEdit(); self.root_edit.setReadOnly(True)
        self.root_edit.setPlaceholderText("Select root data directory…")
        self.root_edit.textChanged.connect(self.params_changed)
        row.addWidget(self.root_edit)
        btn = QPushButton("Browse"); btn.setFixedWidth(72)
        btn.clicked.connect(self._browse_root); row.addWidget(btn)
        v1.addLayout(row); lay.addWidget(g1)

        g2 = QGroupBox("STEP 2  ·  PATIENT INFORMATION")
        v2 = QVBoxLayout(g2); v2.setSpacing(6)
        self.uhid_edit    = self._field(v2, "UHID *")
        self.midas_edit   = self._field(v2, "MIDAS Code / Case_ID *")
        self.curator_edit = self._field(v2, "Curator Name *")
        lbl = QLabel("Visit Date *")
        lbl.setStyleSheet("color:#8B949E;font-size:11px;background:transparent;")
        v2.addWidget(lbl)
        self.date_edit = QDateEdit(QDate.currentDate())
        self.date_edit.setCalendarPopup(True); self.date_edit.setDisplayFormat("dd-MM-yyyy")
        self.date_edit.dateChanged.connect(self.params_changed)
        v2.addWidget(self.date_edit)
        for w in (self.uhid_edit, self.midas_edit, self.curator_edit):
            w.textChanged.connect(self.params_changed)
        lay.addWidget(g2)

        g3 = QGroupBox("STEP 3  ·  CATEGORY")
        v3 = QVBoxLayout(g3); v3.setSpacing(6)
        self.cat_group = QButtonGroup(self)
        self.cat_radios: Dict[str, QRadioButton] = {}
        for code, label in CATEGORIES:
            rb = QRadioButton(f"  {code}   ·   {label}")
            self.cat_group.addButton(rb); self.cat_radios[code] = rb
            rb.toggled.connect(self._on_cat_changed); v3.addWidget(rb)
        lay.addWidget(g3)

        self.g4 = QGroupBox("STEP 4  ·  SUBCATEGORY")
        v4 = QVBoxLayout(self.g4)
        self.subcat_combo = QComboBox()
        self.subcat_combo.currentTextChanged.connect(self._on_subcat_changed)
        v4.addWidget(self.subcat_combo); lay.addWidget(self.g4); self.g4.hide()

        self.g5 = QGroupBox("STEP 5  ·  BODY SITE")
        v5 = QVBoxLayout(self.g5)
        self.body_selector = BodySiteSelector()
        self.body_selector.changed.connect(self.params_changed)
        v5.addWidget(self.body_selector); lay.addWidget(self.g5); self.g5.hide()

        self.g6 = QGroupBox("STEP 6  ·  MAGNIFICATION")
        v6 = QVBoxLayout(self.g6)
        self.mag_combo = QComboBox()
        self.mag_combo.currentTextChanged.connect(self.params_changed)
        v6.addWidget(self.mag_combo); lay.addWidget(self.g6); self.g6.hide()

        lay.addStretch()

    def _field(self, layout, label) -> QLineEdit:
        l = QLabel(label)
        l.setStyleSheet("color:#8B949E;font-size:11px;background:transparent;")
        layout.addWidget(l)
        le = QLineEdit(); le.setPlaceholderText(f"Enter {label.replace(' *','')}…")
        layout.addWidget(le); return le

    def _browse_root(self):
        folder = QFileDialog.getExistingDirectory(
            self, "Select Root Storage Directory", "", QFileDialog.Option.ShowDirsOnly)
        if folder: self.root_edit.setText(folder)

    def _on_cat_changed(self):
        cat = self.get_category()
        self.g4.hide(); self.g5.hide(); self.g6.hide()
        if cat in ("GM", "SM", "OT"):
            self.g4.show()
            self.subcat_combo.blockSignals(True); self.subcat_combo.clear()
            self.subcat_combo.addItems(
                {"GM": GM_SUBCATEGORIES, "SM": SM_SUBCATEGORIES, "OT": OT_SUBCATEGORIES}[cat])
            self.subcat_combo.blockSignals(False)
            self._on_subcat_changed(self.subcat_combo.currentText())
        self.params_changed.emit()

    def _on_subcat_changed(self, subcat):
        cat = self.get_category(); self.g5.hide(); self.g6.hide()
        if subcat == "HISTOPATH" and cat in ("GM", "SM"):
            self.g5.show(); self.g6.show()
            self.mag_combo.blockSignals(True); self.mag_combo.clear()
            self.mag_combo.addItems(MAGNIFICATIONS); self.mag_combo.blockSignals(False)
        elif subcat == "CYTOLOGY":
            self.g6.show()
            self.mag_combo.blockSignals(True); self.mag_combo.clear()
            self.mag_combo.addItems(CYTOLOGY_MAGS); self.mag_combo.blockSignals(False)
        self.params_changed.emit()

    def get_root(self):         return self.root_edit.text().strip()
    def get_uhid(self):         return self.uhid_edit.text().strip()
    def get_midas_code(self):   return self.midas_edit.text().strip()
    def get_curator(self):      return self.curator_edit.text().strip()
    def get_visit_date(self):   return self.date_edit.date().toString("dd-MM-yyyy")
    def get_category(self):
        for code, rb in self.cat_radios.items():
            if rb.isChecked(): return code
        return None
    def get_subcategory(self):
        return self.subcat_combo.currentText() if not self.g4.isHidden() else None
    def get_body_sites(self):
        return [] if self.g5.isHidden() else self.body_selector.get_sites()
    def get_magnification(self):
        return self.mag_combo.currentText() if not self.g6.isHidden() else None

    def validate(self):
        if not self.get_root():       return False, "Please select a root storage directory."
        if not self.get_midas_code(): return False, "MIDAS Code / Case_ID is required."
        if not self.get_uhid():       return False, "UHID is required."
        if not self.get_curator():    return False, "Curator name is required."
        if not self.get_category():   return False, "Please select a category."
        cat, sub = self.get_category(), self.get_subcategory()
        if sub == "HISTOPATH" and cat in ("GM", "SM"):
            if not self.get_body_sites():
                return False, "Please select at least one Body Site."
        return True, ""


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN APP — SAVE CONTROL BAR
# ═══════════════════════════════════════════════════════════════════════════════

class SaveControlBar(QFrame):
    organise_requested    = pyqtSignal()
    save_csv_requested    = pyqtSignal()
    save_log_requested    = pyqtSignal()
    run_qc_requested      = pyqtSignal()
    run_count_requested   = pyqtSignal()    # NEW

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet("QFrame{background:#161B22;border-top:1px solid #21262D;}")
        self.setFixedHeight(90)
        lay = QHBoxLayout(self); lay.setContentsMargins(16,10,16,10); lay.setSpacing(14)

        left = QVBoxLayout(); left.setSpacing(3)
        pl = QLabel("OUTPUT FILENAME PREVIEW")
        pl.setStyleSheet("color:#484F58;font-size:10px;letter-spacing:1.5px;"
                         "font-weight:700;background:transparent;")
        self.preview_lbl = QLabel("—"); self.preview_lbl.setObjectName("filename_preview")
        self.preview_lbl.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        left.addWidget(pl); left.addWidget(self.preview_lbl)
        lay.addLayout(left, stretch=1)

        self.progress = QProgressBar()
        self.progress.setFixedSize(130,6); self.progress.setTextVisible(False)
        self.progress.setValue(0); self.progress.hide()
        lay.addWidget(self.progress)

        self.unsaved_badge = QLabel("UNSAVED CHANGES")
        self.unsaved_badge.setObjectName("unsaved_badge"); self.unsaved_badge.hide()
        lay.addWidget(self.unsaved_badge)

        self.qc_btn = QPushButton("QC Scan"); self.qc_btn.setObjectName("btn_qc")
        self.qc_btn.setFixedSize(120, 48)
        self.qc_btn.setToolTip("Open QC Scanner — audit all images in root folder")
        self.qc_btn.clicked.connect(self.run_qc_requested)
        lay.addWidget(self.qc_btn)

        self.count_btn = QPushButton("Dataset Count")
        self.count_btn.setObjectName("btn_secondary")
        self.count_btn.setFixedSize(138, 48)
        self.count_btn.setToolTip(
            "Count images per modality across all NN_NNN_P case folders")
        self.count_btn.clicked.connect(self.run_count_requested)
        lay.addWidget(self.count_btn)

        self.organise_btn = QPushButton("Organise Images")
        self.organise_btn.setObjectName("btn_primary")
        self.organise_btn.setFixedSize(190, 48); self.organise_btn.setEnabled(False)
        self.organise_btn.clicked.connect(self.organise_requested)
        lay.addWidget(self.organise_btn)

        self.csv_btn = QPushButton("Save CSV"); self.csv_btn.setObjectName("btn_secondary")
        self.csv_btn.setFixedSize(120, 48); self.csv_btn.setEnabled(False)
        self.csv_btn.clicked.connect(self.save_csv_requested)
        lay.addWidget(self.csv_btn)

        self.log_btn = QPushButton("Save Log"); self.log_btn.setObjectName("btn_warning")
        self.log_btn.setFixedSize(110, 48); self.log_btn.setEnabled(False)
        self.log_btn.clicked.connect(self.save_log_requested)
        lay.addWidget(self.log_btn)

    def set_preview(self, t):        self.preview_lbl.setText(t)
    def set_organise_enabled(self, v): self.organise_btn.setEnabled(v)
    def set_csv_enabled(self, v):    self.csv_btn.setEnabled(v)
    def set_log_enabled(self, v):    self.log_btn.setEnabled(v)
    def set_unsaved(self, v):        self.unsaved_badge.setVisible(v)

    def set_progress(self, value, total):
        if total > 0:
            self.progress.setMaximum(total); self.progress.setValue(value)
            self.progress.show()
        else:
            self.progress.hide()


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN WINDOW
# ═══════════════════════════════════════════════════════════════════════════════

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle(f"{APP_TITLE}  {APP_VERSION}")
        self.setMinimumSize(1100, 720); self.resize(1400, 860)
        self._session           = SessionState()
        self._csv_writer:       Optional[CSVWriter]           = None
        self._logger:           Optional[SessionLogger]       = None
        self._pending_csv_rows: List[dict]                    = []
        self._qc_dialog:        Optional[QCReportDialog]      = None
        self._count_dialog:     Optional[DatasetCountDialog]  = None
        self._setup_ui(); self._connect_signals()

    def _setup_ui(self):
        central = QWidget(); central.setObjectName("central")
        self.setCentralWidget(central)
        main = QVBoxLayout(central); main.setContentsMargins(14,14,14,0); main.setSpacing(10)
        self.splitter = QSplitter(Qt.Orientation.Horizontal)
        self.splitter.setChildrenCollapsible(False)
        self.form_panel    = FormPanel()
        self.preview_panel = ImagePreviewPanel()
        self.splitter.addWidget(self.form_panel)
        self.splitter.addWidget(self.preview_panel)
        self.splitter.setSizes([400, 900])
        main.addWidget(self.splitter, stretch=1)
        self.save_bar = SaveControlBar()
        main.addWidget(self.save_bar)
        self.statusBar().setFixedHeight(26)
        self.statusBar().showMessage(f"Ready  ·  {APP_TITLE} {APP_VERSION}")

    def _connect_signals(self):
        self.form_panel.params_changed.connect(self._refresh_preview)
        self.preview_panel.selection_count_changed.connect(self._update_organise_state)
        self.save_bar.organise_requested.connect(self._organise_images)
        self.save_bar.save_csv_requested.connect(self._save_csv)
        self.save_bar.save_log_requested.connect(self._save_log)
        self.save_bar.run_qc_requested.connect(self._run_qc)
        self.save_bar.run_count_requested.connect(self._run_count)

    def _run_qc(self):
        initial = self.form_panel.get_root()
        if not initial or not Path(initial).is_dir():
            src = self.preview_panel.current_source
            initial = str(src) if src else ""
        if self._qc_dialog is None or not self._qc_dialog.isVisible():
            self._qc_dialog = QCReportDialog(initial_folder=initial, parent=self)
        self._qc_dialog.show()
        self._qc_dialog.raise_()
        self._qc_dialog.activateWindow()

    def _run_count(self):
        """Open the Dataset Counter dialog."""
        initial = self.form_panel.get_root()
        if not initial or not Path(initial).is_dir():
            src = self.preview_panel.current_source
            initial = str(src) if src else ""
        if self._count_dialog is None or not self._count_dialog.isVisible():
            self._count_dialog = DatasetCountDialog(
                initial_folder=initial, parent=self)
        self._count_dialog.show()
        self._count_dialog.raise_()
        self._count_dialog.activateWindow()

    def _ensure_logger(self):
        root = self.form_panel.get_root()
        if self._logger is None and root:
            self._logger = SessionLogger(Path(root))
            self.save_bar.set_log_enabled(True)

    def _ensure_csv_writer(self, mc, root):
        midas_folder = Path(root) / mc
        if (self._csv_writer is None or
                self._csv_writer.csv_path.parent != midas_folder):
            self._csv_writer = CSVWriter(midas_folder, mc)

    def _refresh_preview(self):
        mc    = self.form_panel.get_midas_code() or "CASE_ID"
        vd    = self.form_panel.get_visit_date()
        cat   = self.form_panel.get_category() or "XX"
        sites = self.form_panel.get_body_sites()
        bs    = sites[0] if sites else None
        mag   = self.form_panel.get_magnification()
        self.save_bar.set_preview(FileNamer.build(mc, vd, cat, bs, mag, 1))
        self._update_organise_state(
            sum(1 for t in self.preview_panel.thumbnails if t.is_selected))

    def _update_organise_state(self, n):
        ok, _ = self.form_panel.validate()
        self.save_bar.set_organise_enabled(ok and n > 0)

    def _update_unsaved_ui(self):
        self.save_bar.set_unsaved(self._session.has_unsaved or bool(self._pending_csv_rows))
        self.save_bar.set_csv_enabled(bool(self._pending_csv_rows))

    def _organise_images(self):
        try:
            self._do_organise()
        except Exception as exc:
            tb = traceback.format_exc()
            print(tb, file=sys.stderr)
            QMessageBox.critical(self, "Organise Failed",
                                 f"<b>Error:</b> {exc}<br><br><pre>{tb}</pre>")
            self.save_bar.set_progress(0, 0)

    def _do_organise(self):
        ok, msg = self.form_panel.validate()
        if not ok: QMessageBox.warning(self, "Validation Error", msg); return

        selected = self.preview_panel.get_selected_paths()
        if not selected:
            QMessageBox.information(self, "Nothing Selected",
                                    "Please select at least one image."); return

        root    = self.form_panel.get_root()
        mc      = self.form_panel.get_midas_code()
        vd      = self.form_panel.get_visit_date()
        cat     = self.form_panel.get_category()
        subcat  = self.form_panel.get_subcategory()
        sites   = self.form_panel.get_body_sites() or [None]
        mag     = self.form_panel.get_magnification()
        curator = self.form_panel.get_curator()
        uhid    = self.form_panel.get_uhid()

        if not Path(root).exists():
            QMessageBox.warning(self, "Root Path Error",
                                f"Root directory not accessible:\n{root}"); return

        self._ensure_logger(); self._ensure_csv_writer(mc, root)

        total, all_names = 0, []
        self.save_bar.set_progress(0, len(selected) * len(sites))
        step = 0

        for site in sites:
            dest  = FolderBuilder.get_path(root, mc, vd, cat, subcat, site, mag)
            FolderBuilder.ensure(dest)
            start = CounterManager.next_count(dest)
            names: List[str] = []
            for i, src in enumerate(selected):
                if not src.exists():
                    raise FileNotFoundError(f"Source image no longer exists:\n{src}")
                ext      = src.suffix.lower() if src.suffix.lower() in ALL_IMAGE_EXT else ".jpg"
                new_name = FileNamer.build(mc, vd, cat, site, mag, start + i, ext)
                shutil.copy2(str(src), str(dest / new_name))
                names.append(new_name); step += 1
                self.save_bar.set_progress(step, len(selected) * len(sites))
                QApplication.processEvents()

            total += len(names); all_names += names

            row = {h: 0 for h in CSV_HEADERS}
            row.update({"UHID": uhid, "MIDAS_CODE": mc, "VisitDate": vd,
                        "BodySite": site or "", "Curator": curator})
            n = len(names)
            if   cat == "XC": row["XC"] = n
            elif cat == "RG": row["RG"] = n
            elif cat in ("GM", "SM"):
                if subcat == "HISTOPATH":
                    mk = f"Histopath_{mag}" if mag else None
                    if mk and mk in row: row[mk] = n
                elif subcat == "IHC":            row["IHC"] = n
                elif subcat == "SPECIAL_STAINS": row["Special_Stains"] = n
                elif subcat == "CYTOLOGY":       row["Cytology"] = n
            elif cat == "OT":
                if   subcat == "GROSS":   row["Gross"] = n
                elif subcat == "GENOMIC": row["Genomic"] = n
            self._pending_csv_rows.append(row)

            trail = cat + (f" > {subcat}" if subcat else "") + \
                    (f" > {site}" if site else "") + (f" > {mag}" if mag else "")
            if self._logger:
                self._logger.log(
                    f"ORGANISED | Case_ID:{mc} UHID:{uhid} | {n} file(s) | "
                    f"{trail} | Curator:{curator} | Dest:{dest}")

        self.preview_panel.mark_selected_organised()
        self._session.mark_organised(
            f"{total} image(s) — {mc} · {cat}" + (f" / {subcat}" if subcat else ""))
        self._update_unsaved_ui()
        self.statusBar().showMessage(
            f"  {total} image(s) organised  ·  {datetime.now():%H:%M:%S}  ·  CSV not yet saved")
        QMessageBox.information(self, "Images Organised",
                                f"  {total} image(s) organised.\n\n"
                                f"First file:  {all_names[0] if all_names else '—'}\n\n"
                                "Click  'Save CSV'  before closing.")
        QTimer.singleShot(3000, lambda: self.save_bar.set_progress(0, 0))

    @staticmethod
    def _aggregate_rows(rows):
        NUMERIC = {"XC","RG","Gross","Special_Stains","IHC","Cytology","Genomic",
                   "Histopath_4x","Histopath_10x","Histopath_20x",
                   "Histopath_40x","Histopath_100x","WSI"}
        merged: Dict[Tuple[str,str], dict] = {}
        for row in rows:
            key = (str(row.get("MIDAS_CODE","")), str(row.get("VisitDate","")))
            if key not in merged:
                merged[key] = {h: (0 if h in NUMERIC else "") for h in CSV_HEADERS}
            m = merged[key]
            for col in CSV_HEADERS:
                if col in NUMERIC:
                    try: m[col] = int(m[col] or 0) + int(row.get(col) or 0)
                    except (ValueError, TypeError): pass
                else:
                    if not m[col]: m[col] = row.get(col, "")
        return list(merged.values())

    def _save_csv(self):
        if not self._pending_csv_rows:
            QMessageBox.information(self, "Nothing to Save", "No pending rows."); return
        if self._csv_writer is None:
            QMessageBox.warning(self, "CSV Writer Error",
                                "Please organise at least one image first."); return
        try:
            aggregated = self._aggregate_rows(self._pending_csv_rows)
            for row in aggregated:
                self._csv_writer.upsert(row)
            n_p = len(self._pending_csv_rows); n_r = len(aggregated)
            self._pending_csv_rows.clear(); self._session.mark_flushed()
            self._update_unsaved_ui()
            if self._logger:
                self._logger.log(f"CSV SAVED | {n_p} pass(es) -> {n_r} row(s)")
            self.statusBar().showMessage(
                f"  CSV saved  ·  {n_r} row(s)  ->  {self._csv_writer.csv_path}")
            QMessageBox.information(self, "CSV Saved",
                                    f"  {n_r} row(s) written:\n{self._csv_writer.csv_path}")
        except Exception as exc:
            QMessageBox.critical(self, "CSV Save Failed", f"<b>Failed:</b> {exc}")

    def _save_log(self):
        if self._logger is None:
            QMessageBox.information(self, "No Log", "No log created yet."); return
        self._logger.log("USER CONFIRMED LOG SAVE")
        self.statusBar().showMessage(f"  Log: {self._logger.log_path}")
        QMessageBox.information(self, "Log File", f"Session log:\n\n{self._logger.log_path}")

    def closeEvent(self, event: QCloseEvent):
        has_unsaved = self._session.has_unsaved or bool(self._pending_csv_rows)
        if not has_unsaved:
            if self._logger: self._logger.log("=== Session ended ===")
            event.accept(); return

        items = []
        if self._pending_csv_rows:
            items.append(f"  • CSV ({len(self._pending_csv_rows)} unsaved row(s))")
        if self._session.has_unsaved:
            items.append("  • Organised image session records")

        dlg = QMessageBox(self)
        dlg.setWindowTitle("Unsaved Data"); dlg.setIcon(QMessageBox.Icon.Warning)
        dlg.setText("<b>You have unsaved data from this session.</b>")
        dlg.setInformativeText("Unsaved:\n\n" + "\n".join(items) + "\n\nSave before exiting?")
        dlg.setStandardButtons(
            QMessageBox.StandardButton.Save |
            QMessageBox.StandardButton.Discard |
            QMessageBox.StandardButton.Cancel)
        dlg.setDefaultButton(QMessageBox.StandardButton.Save)
        dlg.button(QMessageBox.StandardButton.Save).setText("Save & Exit")
        dlg.button(QMessageBox.StandardButton.Discard).setText("Exit Without Saving")
        result = dlg.exec()

        if result == QMessageBox.StandardButton.Save:
            self._save_csv()
            if self._logger: self._logger.log("=== Session ended (saved) ===")
            event.accept()
        elif result == QMessageBox.StandardButton.Discard:
            if self._logger: self._logger.log("=== Session ended (discarded) ===")
            event.accept()
        else:
            event.ignore()


# ═══════════════════════════════════════════════════════════════════════════════
# GLOBAL EXCEPTION HOOK
# ═══════════════════════════════════════════════════════════════════════════════

def _global_exception_hook(exc_type, exc_value, exc_tb):
    tb_text = "".join(traceback.format_exception(exc_type, exc_value, exc_tb))
    print(tb_text, file=sys.stderr)
    app = QApplication.instance()
    if app:
        msg = QMessageBox()
        msg.setWindowTitle("Unexpected Error"); msg.setIcon(QMessageBox.Icon.Critical)
        msg.setText("<b>An unexpected error occurred.</b>")
        msg.setDetailedText(tb_text)
        msg.setStandardButtons(QMessageBox.StandardButton.Ok)
        msg.exec()


def main():
    app = QApplication(sys.argv)
    app.setApplicationName(APP_TITLE)
    app.setApplicationVersion(APP_VERSION)
    app.setStyleSheet(QSS)
    sys.excepthook = _global_exception_hook
    window = MainWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
